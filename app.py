
import os
import random
import re
import sqlite3
import string
import tempfile
import time
from datetime import date, datetime, timedelta
from functools import wraps
from io import BytesIO
from urllib.parse import urlparse

from flask import (
    Flask,
    abort,
    flash,
    g,
    redirect,
    render_template,
    request,
    send_file,
    session,
    url_for,
)
from werkzeug.security import check_password_hash, generate_password_hash

try:
    from openpyxl import Workbook
except ImportError:
    Workbook = None


BASE_DIR = os.path.abspath(os.path.dirname(__file__))
DATA_DIR = os.path.join(BASE_DIR, "data")
DB_PATH = os.environ.get("DATABASE_PATH", os.path.join(DATA_DIR, "commission.db"))
MANUAL_PDF_PATH = os.environ.get(
    "MANUAL_PDF_PATH",
    os.path.join(BASE_DIR, "static", "manual", "Manual_do_Usuario_Dashboard_de_Vendas.pdf"),
)

PAYMENT_FORMATS = {
    "avista": "À vista",
    "parcelado": "Parcelado",
    "recorrencia": "Recorrência",
}
COMMISSION_PAYMENT_MODES = {
    "per_installment": "Comissão em recorrência (por parcela)",
    "upfront_first_installment": "Comissão total na 1ª parcela",
}
INSTALLMENT_STATUSES = {
    "confirmado": "Confirmado",
    "cancelado": "Cancelado",
    "atrasado": "Atrasado",
}
ROLES = ("admin", "seller", "viewer")
ROLE_LABELS = {"admin": "Administrador", "seller": "Vendedor", "viewer": "Visualizador"}
REQUEST_STATUSES = ("pendente", "aprovado", "recusado")
JOIN_REQUEST_PROFILES = ("seller", "viewer")
MASTER_USERNAME = "marcosmello2402"
BUILTIN_MASTER_USERNAME = "admin"
BUILTIN_MASTER_PASSWORD = "123"
BUILTIN_MASTER_FULL_NAME = "Administrador Mestre"

SCHEMA_SQL = """
PRAGMA foreign_keys = ON;

CREATE TABLE IF NOT EXISTS users (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    username TEXT NOT NULL UNIQUE,
    email TEXT,
    full_name TEXT NOT NULL,
    role TEXT NOT NULL CHECK (role IN ('admin', 'seller', 'viewer')),
    password_hash TEXT NOT NULL,
    password_plaintext TEXT,
    header_label TEXT,
    invited_by_username TEXT,
    invited_by_email TEXT,
    owner_admin_id INTEGER,
    is_master INTEGER NOT NULL DEFAULT 0,
    is_manager INTEGER NOT NULL DEFAULT 0,
    is_active INTEGER NOT NULL DEFAULT 1,
    created_at TEXT NOT NULL,
    FOREIGN KEY (owner_admin_id) REFERENCES users(id) ON DELETE SET NULL
);

CREATE TABLE IF NOT EXISTS communities (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    name TEXT NOT NULL UNIQUE,
    owner_admin_id INTEGER,
    manager_user_id INTEGER,
    is_active INTEGER NOT NULL DEFAULT 1,
    created_at TEXT NOT NULL,
    updated_at TEXT NOT NULL,
    FOREIGN KEY (owner_admin_id) REFERENCES users(id) ON DELETE SET NULL,
    FOREIGN KEY (manager_user_id) REFERENCES users(id) ON DELETE SET NULL
);

CREATE TABLE IF NOT EXISTS companies (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    name TEXT NOT NULL UNIQUE,
    owner_admin_id INTEGER,
    is_active INTEGER NOT NULL DEFAULT 1,
    created_at TEXT NOT NULL,
    updated_at TEXT NOT NULL,
    FOREIGN KEY (owner_admin_id) REFERENCES users(id) ON DELETE SET NULL
);

CREATE TABLE IF NOT EXISTS courses (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    company_id INTEGER NOT NULL,
    name TEXT NOT NULL,
    default_commission_percent REAL NOT NULL CHECK (default_commission_percent >= 0 AND default_commission_percent <= 100),
    is_active INTEGER NOT NULL DEFAULT 1,
    created_at TEXT NOT NULL,
    updated_at TEXT NOT NULL,
    UNIQUE (company_id, name),
    FOREIGN KEY (company_id) REFERENCES companies(id) ON DELETE RESTRICT
);

CREATE TABLE IF NOT EXISTS sales (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    sale_date TEXT NOT NULL,
    customer_name TEXT NOT NULL,
    customer_phone TEXT,
    customer_email TEXT,
    company_id INTEGER NOT NULL,
    course_id INTEGER NOT NULL,
    seller_id INTEGER NOT NULL,
    payment_format TEXT NOT NULL CHECK (payment_format IN ('avista', 'parcelado', 'recorrencia')),
    commission_payment_mode TEXT NOT NULL DEFAULT 'per_installment',
    installments_count INTEGER NOT NULL CHECK (installments_count >= 1),
    total_value REAL NOT NULL CHECK (total_value > 0),
    commission_percent REAL NOT NULL CHECK (commission_percent >= 0 AND commission_percent <= 100),
    total_commission_expected REAL NOT NULL,
    owner_admin_id INTEGER,
    notes TEXT,
    created_by INTEGER NOT NULL,
    created_at TEXT NOT NULL,
    updated_at TEXT NOT NULL,
    FOREIGN KEY (company_id) REFERENCES companies(id) ON DELETE RESTRICT,
    FOREIGN KEY (course_id) REFERENCES courses(id) ON DELETE RESTRICT,
    FOREIGN KEY (seller_id) REFERENCES users(id) ON DELETE RESTRICT,
    FOREIGN KEY (created_by) REFERENCES users(id) ON DELETE RESTRICT,
    FOREIGN KEY (owner_admin_id) REFERENCES users(id) ON DELETE SET NULL
);

CREATE TABLE IF NOT EXISTS sale_installments (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    sale_id INTEGER NOT NULL,
    installment_number INTEGER NOT NULL,
    due_date TEXT NOT NULL,
    month_key TEXT NOT NULL,
    installment_value REAL NOT NULL,
    commission_value REAL NOT NULL,
    status TEXT NOT NULL CHECK (status IN ('confirmado', 'cancelado', 'atrasado')),
    paid_at TEXT,
    seller_followup_checked INTEGER NOT NULL DEFAULT 0,
    seller_followup_checked_at TEXT,
    created_at TEXT NOT NULL,
    updated_at TEXT NOT NULL,
    UNIQUE (sale_id, installment_number),
    FOREIGN KEY (sale_id) REFERENCES sales(id) ON DELETE CASCADE
);

CREATE TABLE IF NOT EXISTS viewer_course_requests (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    viewer_id INTEGER NOT NULL,
    course_id INTEGER NOT NULL,
    request_note TEXT,
    status TEXT NOT NULL CHECK (status IN ('pendente', 'aprovado', 'recusado')),
    requested_at TEXT NOT NULL,
    reviewed_at TEXT,
    reviewed_by INTEGER,
    approval_days INTEGER,
    is_permanent INTEGER NOT NULL DEFAULT 0,
    UNIQUE (viewer_id, course_id, status),
    FOREIGN KEY (viewer_id) REFERENCES users(id) ON DELETE CASCADE,
    FOREIGN KEY (course_id) REFERENCES courses(id) ON DELETE CASCADE,
    FOREIGN KEY (reviewed_by) REFERENCES users(id) ON DELETE SET NULL
);

CREATE TABLE IF NOT EXISTS viewer_course_access (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    viewer_id INTEGER NOT NULL,
    course_id INTEGER NOT NULL,
    granted_by INTEGER NOT NULL,
    granted_at TEXT NOT NULL,
    expires_at TEXT,
    is_permanent INTEGER NOT NULL DEFAULT 0,
    UNIQUE (viewer_id, course_id),
    FOREIGN KEY (viewer_id) REFERENCES users(id) ON DELETE CASCADE,
    FOREIGN KEY (course_id) REFERENCES courses(id) ON DELETE CASCADE,
    FOREIGN KEY (granted_by) REFERENCES users(id) ON DELETE RESTRICT
);

CREATE TABLE IF NOT EXISTS user_access_requests (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    requested_username TEXT NOT NULL,
    requested_full_name TEXT NOT NULL,
    desired_profile TEXT NOT NULL CHECK (desired_profile IN ('manager', 'seller', 'viewer')),
    password_hash TEXT NOT NULL,
    password_plaintext TEXT,
    referral_username TEXT NOT NULL,
    owner_admin_id INTEGER,
    requested_community_id INTEGER,
    status TEXT NOT NULL CHECK (status IN ('pendente', 'aprovado', 'recusado')),
    request_note TEXT,
    requested_at TEXT NOT NULL,
    reviewed_at TEXT,
    reviewed_by INTEGER,
    decision_note TEXT,
    FOREIGN KEY (owner_admin_id) REFERENCES users(id) ON DELETE SET NULL,
    FOREIGN KEY (requested_community_id) REFERENCES communities(id) ON DELETE SET NULL,
    FOREIGN KEY (reviewed_by) REFERENCES users(id) ON DELETE SET NULL
);

CREATE TABLE IF NOT EXISTS user_community_memberships (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    user_id INTEGER NOT NULL,
    owner_admin_id INTEGER NOT NULL,
    is_active INTEGER NOT NULL DEFAULT 1,
    created_at TEXT NOT NULL,
    updated_at TEXT NOT NULL,
    UNIQUE (user_id, owner_admin_id),
    FOREIGN KEY (user_id) REFERENCES users(id) ON DELETE CASCADE,
    FOREIGN KEY (owner_admin_id) REFERENCES users(id) ON DELETE CASCADE
);

CREATE TABLE IF NOT EXISTS seller_course_permissions (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    user_id INTEGER NOT NULL,
    course_id INTEGER NOT NULL,
    can_launch_sales INTEGER NOT NULL DEFAULT 1,
    can_edit_sales INTEGER NOT NULL DEFAULT 1,
    can_edit_course INTEGER NOT NULL DEFAULT 0,
    is_active INTEGER NOT NULL DEFAULT 1,
    created_at TEXT NOT NULL,
    updated_at TEXT NOT NULL,
    UNIQUE (user_id, course_id),
    FOREIGN KEY (user_id) REFERENCES users(id) ON DELETE CASCADE,
    FOREIGN KEY (course_id) REFERENCES courses(id) ON DELETE CASCADE
);

CREATE INDEX IF NOT EXISTS idx_sales_sale_date ON sales(sale_date);
CREATE INDEX IF NOT EXISTS idx_sales_company ON sales(company_id);
CREATE INDEX IF NOT EXISTS idx_sales_course ON sales(course_id);
CREATE INDEX IF NOT EXISTS idx_sales_seller ON sales(seller_id);

CREATE INDEX IF NOT EXISTS idx_installments_sale_id ON sale_installments(sale_id);
CREATE INDEX IF NOT EXISTS idx_installments_due_date ON sale_installments(due_date);
CREATE INDEX IF NOT EXISTS idx_installments_month_key ON sale_installments(month_key);
CREATE INDEX IF NOT EXISTS idx_installments_status ON sale_installments(status);
CREATE INDEX IF NOT EXISTS idx_requests_viewer ON viewer_course_requests(viewer_id);
CREATE INDEX IF NOT EXISTS idx_requests_course ON viewer_course_requests(course_id);
CREATE INDEX IF NOT EXISTS idx_access_viewer ON viewer_course_access(viewer_id);
CREATE INDEX IF NOT EXISTS idx_access_course ON viewer_course_access(course_id);
CREATE INDEX IF NOT EXISTS idx_communities_owner_admin ON communities(owner_admin_id);
CREATE INDEX IF NOT EXISTS idx_requests_owner_admin ON user_access_requests(owner_admin_id);
CREATE INDEX IF NOT EXISTS idx_requests_status ON user_access_requests(status);
CREATE INDEX IF NOT EXISTS idx_memberships_user ON user_community_memberships(user_id);
CREATE INDEX IF NOT EXISTS idx_memberships_owner ON user_community_memberships(owner_admin_id);
CREATE INDEX IF NOT EXISTS idx_seller_perm_user ON seller_course_permissions(user_id);
CREATE INDEX IF NOT EXISTS idx_seller_perm_course ON seller_course_permissions(course_id);
"""


app = Flask(__name__)
app.config["SECRET_KEY"] = os.environ.get("SECRET_KEY", "change-this-secret-in-production")
app.config["DATABASE"] = DB_PATH
app.config["SESSION_COOKIE_HTTPONLY"] = True
app.config["SESSION_COOKIE_SAMESITE"] = "Lax"


def database_needs_reset(db_path):
    if not os.path.exists(db_path):
        return False
    try:
        conn = sqlite3.connect(db_path)
        # quick_check força leitura real do arquivo SQLite e detecta corrupção/trava de I/O.
        check = conn.execute("PRAGMA quick_check").fetchone()
        conn.close()
        if check and str(check[0]).lower() != "ok":
            return True
        return False
    except sqlite3.Error:
        return True


def quarantine_broken_db(db_path):
    if not os.path.exists(db_path):
        return
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    db_dir = os.path.dirname(db_path) or DATA_DIR
    db_name = os.path.splitext(os.path.basename(db_path))[0]
    os.makedirs(db_dir, exist_ok=True)

    backup_db = os.path.join(db_dir, f"{db_name}_legacy_{timestamp}.db")
    try:
        os.replace(db_path, backup_db)
    except OSError:
        pass

    for suffix in ("-journal", "-wal", "-shm"):
        sidecar = f"{db_path}{suffix}"
        if os.path.exists(sidecar):
            try:
                os.remove(sidecar)
            except OSError:
                pass


def ensure_column(conn, table_name, column_name, definition_sql):
    cols = {row[1] for row in conn.execute(f"PRAGMA table_info({table_name})").fetchall()}
    if column_name not in cols:
        conn.execute(f"ALTER TABLE {table_name} ADD COLUMN {definition_sql}")


def rebuild_communities_table_if_needed(conn):
    row = conn.execute(
        "SELECT sql FROM sqlite_master WHERE type = 'table' AND name = 'communities'"
    ).fetchone()
    if not row or not row[0]:
        return
    sql_lower = row[0].lower()
    needs_rebuild = "owner_admin_id integer not null unique" in sql_lower or "owner_admin_id integer not null" in sql_lower
    if not needs_rebuild:
        return

    conn.execute("ALTER TABLE communities RENAME TO communities_legacy")
    conn.execute(
        """
        CREATE TABLE communities (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL UNIQUE,
            owner_admin_id INTEGER,
            manager_user_id INTEGER,
            is_active INTEGER NOT NULL DEFAULT 1,
            created_at TEXT NOT NULL,
            updated_at TEXT NOT NULL,
            FOREIGN KEY (owner_admin_id) REFERENCES users(id) ON DELETE SET NULL,
            FOREIGN KEY (manager_user_id) REFERENCES users(id) ON DELETE SET NULL
        )
        """
    )
    conn.execute(
        """
        INSERT INTO communities (id, name, owner_admin_id, manager_user_id, is_active, created_at, updated_at)
        SELECT id, name, owner_admin_id, manager_user_id, is_active, created_at, updated_at
        FROM communities_legacy
        """
    )
    conn.execute("DROP TABLE communities_legacy")


def rebuild_user_access_requests_table_if_needed(conn):
    row = conn.execute(
        "SELECT sql FROM sqlite_master WHERE type = 'table' AND name = 'user_access_requests'"
    ).fetchone()
    if not row or not row[0]:
        return
    sql_lower = row[0].lower()
    needs_rebuild = (
        "owner_admin_id integer not null" in sql_lower
        or "password_plaintext" not in sql_lower
        or "requested_community_id" not in sql_lower
    )
    if not needs_rebuild:
        return

    conn.execute("ALTER TABLE user_access_requests RENAME TO user_access_requests_legacy")
    conn.execute(
        """
        CREATE TABLE user_access_requests (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            requested_username TEXT NOT NULL,
            requested_full_name TEXT NOT NULL,
            desired_profile TEXT NOT NULL CHECK (desired_profile IN ('manager', 'seller', 'viewer')),
            password_hash TEXT NOT NULL,
            password_plaintext TEXT,
            referral_username TEXT NOT NULL,
            owner_admin_id INTEGER,
            requested_community_id INTEGER,
            status TEXT NOT NULL CHECK (status IN ('pendente', 'aprovado', 'recusado')),
            request_note TEXT,
            requested_at TEXT NOT NULL,
            reviewed_at TEXT,
            reviewed_by INTEGER,
            decision_note TEXT,
            FOREIGN KEY (owner_admin_id) REFERENCES users(id) ON DELETE SET NULL,
            FOREIGN KEY (requested_community_id) REFERENCES communities(id) ON DELETE SET NULL,
            FOREIGN KEY (reviewed_by) REFERENCES users(id) ON DELETE SET NULL
        )
        """
    )
    conn.execute(
        """
        INSERT INTO user_access_requests (
            id,
            requested_username,
            requested_full_name,
            desired_profile,
            password_hash,
            password_plaintext,
            referral_username,
            owner_admin_id,
            requested_community_id,
            status,
            request_note,
            requested_at,
            reviewed_at,
            reviewed_by,
            decision_note
        )
        SELECT
            id,
            requested_username,
            requested_full_name,
            desired_profile,
            password_hash,
            NULL,
            referral_username,
            owner_admin_id,
            NULL,
            status,
            request_note,
            requested_at,
            reviewed_at,
            reviewed_by,
            decision_note
        FROM user_access_requests_legacy
        """
    )
    conn.execute("DROP TABLE user_access_requests_legacy")


def apply_schema_migrations(conn):
    ensure_column(conn, "users", "email", "email TEXT")
    ensure_column(conn, "users", "invited_by_username", "invited_by_username TEXT")
    ensure_column(conn, "users", "invited_by_email", "invited_by_email TEXT")
    ensure_column(conn, "users", "owner_admin_id", "owner_admin_id INTEGER")
    ensure_column(conn, "users", "is_master", "is_master INTEGER NOT NULL DEFAULT 0")
    ensure_column(conn, "users", "is_manager", "is_manager INTEGER NOT NULL DEFAULT 0")
    ensure_column(conn, "users", "password_plaintext", "password_plaintext TEXT")

    ensure_column(conn, "companies", "owner_admin_id", "owner_admin_id INTEGER")

    ensure_column(conn, "sales", "owner_admin_id", "owner_admin_id INTEGER")
    ensure_column(
        conn,
        "sales",
        "commission_payment_mode",
        "commission_payment_mode TEXT NOT NULL DEFAULT 'per_installment'",
    )

    ensure_column(conn, "sale_installments", "seller_followup_checked", "seller_followup_checked INTEGER NOT NULL DEFAULT 0")
    ensure_column(conn, "sale_installments", "seller_followup_checked_at", "seller_followup_checked_at TEXT")

    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS communities (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL UNIQUE,
            owner_admin_id INTEGER,
            manager_user_id INTEGER,
            is_active INTEGER NOT NULL DEFAULT 1,
            created_at TEXT NOT NULL,
            updated_at TEXT NOT NULL,
            FOREIGN KEY (owner_admin_id) REFERENCES users(id) ON DELETE SET NULL,
            FOREIGN KEY (manager_user_id) REFERENCES users(id) ON DELETE SET NULL
        )
        """
    )
    rebuild_communities_table_if_needed(conn)
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS user_access_requests (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            requested_username TEXT NOT NULL,
            requested_full_name TEXT NOT NULL,
            desired_profile TEXT NOT NULL CHECK (desired_profile IN ('manager', 'seller', 'viewer')),
            password_hash TEXT NOT NULL,
            password_plaintext TEXT,
            referral_username TEXT NOT NULL,
            owner_admin_id INTEGER,
            requested_community_id INTEGER,
            status TEXT NOT NULL CHECK (status IN ('pendente', 'aprovado', 'recusado')),
            request_note TEXT,
            requested_at TEXT NOT NULL,
            reviewed_at TEXT,
            reviewed_by INTEGER,
            decision_note TEXT,
            FOREIGN KEY (owner_admin_id) REFERENCES users(id) ON DELETE SET NULL,
            FOREIGN KEY (requested_community_id) REFERENCES communities(id) ON DELETE SET NULL,
            FOREIGN KEY (reviewed_by) REFERENCES users(id) ON DELETE SET NULL
        )
        """
    )
    rebuild_user_access_requests_table_if_needed(conn)
    ensure_column(conn, "user_access_requests", "password_plaintext", "password_plaintext TEXT")
    ensure_column(conn, "user_access_requests", "requested_community_id", "requested_community_id INTEGER")
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS user_community_memberships (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            owner_admin_id INTEGER NOT NULL,
            is_active INTEGER NOT NULL DEFAULT 1,
            created_at TEXT NOT NULL,
            updated_at TEXT NOT NULL,
            UNIQUE (user_id, owner_admin_id),
            FOREIGN KEY (user_id) REFERENCES users(id) ON DELETE CASCADE,
            FOREIGN KEY (owner_admin_id) REFERENCES users(id) ON DELETE CASCADE
        )
        """
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS seller_course_permissions (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            course_id INTEGER NOT NULL,
            can_launch_sales INTEGER NOT NULL DEFAULT 1,
            can_edit_sales INTEGER NOT NULL DEFAULT 1,
            can_edit_course INTEGER NOT NULL DEFAULT 0,
            is_active INTEGER NOT NULL DEFAULT 1,
            created_at TEXT NOT NULL,
            updated_at TEXT NOT NULL,
            UNIQUE (user_id, course_id),
            FOREIGN KEY (user_id) REFERENCES users(id) ON DELETE CASCADE,
            FOREIGN KEY (course_id) REFERENCES courses(id) ON DELETE CASCADE
        )
        """
    )

    conn.execute("CREATE INDEX IF NOT EXISTS idx_users_owner_admin ON users(owner_admin_id)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_companies_owner_admin ON companies(owner_admin_id)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_sales_owner_admin ON sales(owner_admin_id)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_communities_owner_admin ON communities(owner_admin_id)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_requests_owner_admin ON user_access_requests(owner_admin_id)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_requests_community ON user_access_requests(requested_community_id)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_requests_status ON user_access_requests(status)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_memberships_user ON user_community_memberships(user_id)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_memberships_owner ON user_community_memberships(owner_admin_id)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_seller_perm_user ON seller_course_permissions(user_id)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_seller_perm_course ON seller_course_permissions(course_id)")


def bootstrap_multitenant_data(conn):
    conn.row_factory = sqlite3.Row
    users = conn.execute(
        """
        SELECT
          id,
          username,
          email,
          role,
          owner_admin_id,
          invited_by_username,
          invited_by_email,
          is_master,
          is_manager
        FROM users
        ORDER BY id ASC
        """
    ).fetchall()
    if not users:
        return

    conn.execute(
        """
        UPDATE users
        SET is_master = 1,
            role = 'admin',
            is_manager = 0
        WHERE lower(username) = ?
        """,
        (MASTER_USERNAME,),
    )
    conn.execute("UPDATE users SET is_manager = COALESCE(is_manager, 0)")
    master_total = conn.execute(
        """
        SELECT COUNT(*) AS total
        FROM users
        WHERE role = 'admin' AND COALESCE(is_master, 0) = 1
        """
    ).fetchone()["total"]
    if master_total == 0:
        fallback_master = conn.execute(
            """
            SELECT id
            FROM users
            WHERE role = 'admin'
            ORDER BY id ASC
            LIMIT 1
            """
        ).fetchone()
        if fallback_master:
            conn.execute("UPDATE users SET is_master = 1 WHERE id = ?", (fallback_master["id"],))

    conn.execute(
        """
        UPDATE users
        SET owner_admin_id = id
        WHERE role = 'admin'
          AND COALESCE(is_master, 0) = 0
          AND (owner_admin_id IS NULL OR owner_admin_id = 0)
        """
    )
    conn.execute("UPDATE users SET owner_admin_id = NULL WHERE COALESCE(is_master, 0) = 1")

    admins = conn.execute(
        """
        SELECT id, username, email
        FROM users
        WHERE role = 'admin'
          AND COALESCE(is_master, 0) = 0
        ORDER BY id ASC
        """
    ).fetchall()
    first_admin = admins[0] if admins else None

    for admin in admins:
        admin_username = (admin["username"] or "").strip().lower()
        if not admin_username:
            continue
        conn.execute(
            """
            UPDATE users
            SET owner_admin_id = ?
            WHERE role <> 'admin'
              AND (owner_admin_id IS NULL OR owner_admin_id = 0)
              AND lower(COALESCE(invited_by_username, '')) = lower(?)
            """,
            (admin["id"], admin_username),
        )
        if admin["email"]:
            conn.execute(
                """
                UPDATE users
                SET owner_admin_id = ?,
                    invited_by_username = COALESCE(NULLIF(invited_by_username, ''), ?)
                WHERE role <> 'admin'
                  AND (owner_admin_id IS NULL OR owner_admin_id = 0)
                  AND lower(COALESCE(invited_by_email, '')) = lower(?)
                """,
                (admin["id"], admin_username, admin["email"]),
            )

    if first_admin:
        conn.execute(
            """
            UPDATE users
            SET owner_admin_id = ?,
                invited_by_username = COALESCE(NULLIF(invited_by_username, ''), ?)
            WHERE role <> 'admin'
              AND (owner_admin_id IS NULL OR owner_admin_id = 0)
            """,
            (first_admin["id"], first_admin["username"]),
        )
        conn.execute(
            "UPDATE companies SET owner_admin_id = ? WHERE owner_admin_id IS NULL",
            (first_admin["id"],),
        )
    else:
        # Master-only bootstrap: keep data global until a community admin exists.
        conn.execute("UPDATE companies SET owner_admin_id = NULL")

    conn.execute(
        """
        UPDATE sales
        SET owner_admin_id = (
            SELECT COALESCE(NULLIF(u.owner_admin_id, 0), CASE WHEN u.role = 'admin' THEN u.id END)
            FROM users u
            WHERE u.id = sales.seller_id
        )
        WHERE owner_admin_id IS NULL
        """
    )
    conn.execute(
        """
        UPDATE sales
        SET commission_payment_mode = 'per_installment'
        WHERE commission_payment_mode IS NULL OR trim(commission_payment_mode) = ''
        """
    )
    conn.execute(
        """
        UPDATE sales
        SET commission_payment_mode = 'upfront_first_installment'
        WHERE payment_format IN ('avista', 'parcelado')
        """
    )
    conn.execute(
        """
        UPDATE sale_installments
        SET commission_value = CASE
            WHEN installment_number = 1 THEN COALESCE((SELECT total_commission_expected FROM sales s WHERE s.id = sale_installments.sale_id), 0)
            ELSE 0
        END,
            updated_at = ?
        WHERE sale_id IN (
            SELECT id
            FROM sales
            WHERE commission_payment_mode = 'upfront_first_installment'
        )
        """,
        (now_iso(),),
    )
    conn.execute(
        """
        UPDATE sale_installments
        SET seller_followup_checked = COALESCE(seller_followup_checked, 0)
        """
    )
    conn.execute(
        """
        UPDATE users
        SET invited_by_username = (
            SELECT admin.username
            FROM users admin
            WHERE admin.id = users.owner_admin_id
        )
        WHERE role <> 'admin'
          AND (invited_by_username IS NULL OR trim(invited_by_username) = '')
          AND owner_admin_id IS NOT NULL
        """
    )
    # Garante vinculações de comunidade para permitir acesso a múltiplas comunidades.
    conn.execute(
        """
        INSERT INTO user_community_memberships (user_id, owner_admin_id, is_active, created_at, updated_at)
        SELECT u.id, u.owner_admin_id, 1, ?, ?
        FROM users u
        WHERE u.owner_admin_id IS NOT NULL
          AND u.role <> 'admin'
        ON CONFLICT(user_id, owner_admin_id)
        DO UPDATE SET is_active = 1, updated_at = excluded.updated_at
        """,
        (now_iso(), now_iso()),
    )
    conn.execute(
        """
        INSERT INTO user_community_memberships (user_id, owner_admin_id, is_active, created_at, updated_at)
        SELECT u.id, u.id, 1, ?, ?
        FROM users u
        WHERE u.role = 'admin' AND COALESCE(u.is_master, 0) = 0
        ON CONFLICT(user_id, owner_admin_id)
        DO UPDATE SET is_active = 1, updated_at = excluded.updated_at
        """,
        (now_iso(), now_iso()),
    )
    # Compatibilidade: vendedores existentes recebem permissão inicial em cursos da comunidade.
    conn.execute(
        """
        INSERT INTO seller_course_permissions (
            user_id, course_id, can_launch_sales, can_edit_sales, can_edit_course, is_active, created_at, updated_at
        )
        SELECT u.id, c.id, 1, 1, 0, 1, ?, ?
        FROM users u
        INNER JOIN companies comp ON comp.owner_admin_id = u.owner_admin_id
        INNER JOIN courses c ON c.company_id = comp.id
        WHERE u.role = 'seller'
          AND COALESCE(u.owner_admin_id, 0) > 0
        ON CONFLICT(user_id, course_id)
        DO NOTHING
        """,
        (now_iso(), now_iso()),
    )

    # Auto-cria comunidade para cada admin de comunidade ativo.
    community_admins = conn.execute(
        """
        SELECT id, username
        FROM users
        WHERE role = 'admin' AND COALESCE(is_master, 0) = 0
        ORDER BY id ASC
        """
    ).fetchall()
    for admin in community_admins:
        exists = conn.execute("SELECT id FROM communities WHERE owner_admin_id = ?", (admin["id"],)).fetchone()
        if exists:
            continue
        now = now_iso()
        conn.execute(
            """
            INSERT INTO communities (name, owner_admin_id, manager_user_id, is_active, created_at, updated_at)
            VALUES (?, ?, NULL, 1, ?, ?)
            """,
            (f"Comunidade {admin['username']}", admin["id"], now, now),
        )

    # Se houver gestores já marcados, tenta vinculá-los automaticamente à comunidade.
    manager_rows = conn.execute(
        """
        SELECT id, owner_admin_id
        FROM users
        WHERE role = 'seller' AND COALESCE(is_manager, 0) = 1 AND owner_admin_id IS NOT NULL
        ORDER BY id ASC
        """
    ).fetchall()
    for manager in manager_rows:
        conn.execute(
            """
            UPDATE communities
            SET manager_user_id = COALESCE(manager_user_id, ?), updated_at = ?
            WHERE owner_admin_id = ?
            """,
            (manager["id"], now_iso(), manager["owner_admin_id"]),
        )


def initialize_database_file(db_path):
    db_dir = os.path.dirname(db_path) or DATA_DIR
    os.makedirs(db_dir, exist_ok=True)
    os.makedirs(DATA_DIR, exist_ok=True)
    legacy_users = []

    if database_needs_reset(db_path):
        try:
            legacy_users = read_legacy_users(db_path)
        except sqlite3.Error:
            legacy_users = []
        quarantine_broken_db(db_path)

    retryable_errors = ("database is locked", "disk i/o error", "unable to open database file")
    last_error = None
    recovery_attempted = False
    for attempt in range(3):
        conn = sqlite3.connect(db_path, timeout=15)
        try:
            conn.executescript(SCHEMA_SQL)
            apply_schema_migrations(conn)
            restore_users(conn, legacy_users)
            bootstrap_multitenant_data(conn)
            seed_defaults(conn)
            bootstrap_multitenant_data(conn)
            apply_bootstrap_admin_from_env(conn)
            ensure_builtin_master_admin(conn)
            conn.commit()
            return
        except sqlite3.OperationalError as exc:
            last_error = exc
            msg = str(exc).lower()
            if any(token in msg for token in retryable_errors) and not recovery_attempted:
                recovery_attempted = True
                try:
                    conn.close()
                except Exception:
                    pass
                quarantine_broken_db(db_path)
                legacy_users = []
                time.sleep(0.6)
                continue
            if any(token in msg for token in retryable_errors) and attempt < 2:
                time.sleep(0.8 * (attempt + 1))
                continue
            raise
        finally:
            conn.close()
    if last_error:
        raise last_error


def init_db():
    app.config["DATABASE"] = DB_PATH
    try:
        initialize_database_file(DB_PATH)
        return
    except sqlite3.OperationalError as exc:
        msg = str(exc).lower()
        retryable_errors = ("database is locked", "disk i/o error", "unable to open database file")
        if not any(token in msg for token in retryable_errors):
            raise

    fallback_path = os.environ.get(
        "DATABASE_FALLBACK_PATH",
        os.path.join(tempfile.gettempdir(), "dashboard_vendas_fallback.db"),
    )
    print(f"[init_db] Banco principal indisponível ({DB_PATH}). Usando fallback em: {fallback_path}")
    app.config["DATABASE"] = fallback_path
    initialize_database_file(fallback_path)


def read_legacy_users(db_path):
    if not os.path.exists(db_path):
        return []
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    try:
        table = conn.execute(
            "SELECT name FROM sqlite_master WHERE type='table' AND name='users'"
        ).fetchone()
        if not table:
            return []
        rows = conn.execute(
            "SELECT id, username, full_name, role, password_hash, is_active, created_at FROM users"
        ).fetchall()
        return [dict(row) for row in rows]
    finally:
        conn.close()


def restore_users(conn, users):
    if not users:
        return
    existing = conn.execute("SELECT COUNT(*) FROM users").fetchone()[0]
    if existing:
        return
    for user in users:
        conn.execute(
            """
            INSERT INTO users (
                id, username, email, full_name, role, password_hash, header_label,
                invited_by_username, invited_by_email, owner_admin_id, is_master, is_manager, is_active, created_at
            )
            VALUES (?, ?, NULL, ?, ?, ?, NULL, NULL, NULL, NULL, 0, 0, ?, ?)
            """,
            (
                user["id"],
                user["username"],
                user["full_name"],
                user["role"] if user["role"] in ROLES else "viewer",
                user["password_hash"],
                user["is_active"],
                user["created_at"],
            ),
        )


def seed_defaults(conn):
    user_count = conn.execute("SELECT COUNT(*) FROM users").fetchone()[0]
    if user_count == 0:
        return

    company_count = conn.execute("SELECT COUNT(*) FROM companies").fetchone()[0]
    if company_count == 0:
        now = now_iso()
        community_owner = conn.execute(
            """
            SELECT id
            FROM users
            WHERE role = 'admin' AND COALESCE(is_master, 0) = 0
            ORDER BY id ASC
            LIMIT 1
            """
        ).fetchone()
        cursor = conn.execute(
            "INSERT INTO companies (name, owner_admin_id, is_active, created_at, updated_at) VALUES (?, ?, 1, ?, ?)",
            ("Empresa Padrão", community_owner["id"] if community_owner else None, now, now),
        )
        company_id = cursor.lastrowid
        conn.execute(
            """
            INSERT INTO courses (company_id, name, default_commission_percent, is_active, created_at, updated_at)
            VALUES (?, ?, ?, 1, ?, ?)
            """,
            (company_id, "Curso Padrão", 10, now, now),
        )


def apply_bootstrap_admin_from_env(conn):
    enabled = (os.environ.get("BOOTSTRAP_ADMIN_ENABLE", "") or "").strip().lower()
    if enabled not in ("1", "true", "yes", "on"):
        return

    username = normalize_username(os.environ.get("BOOTSTRAP_ADMIN_USERNAME"))
    password = os.environ.get("BOOTSTRAP_ADMIN_PASSWORD") or ""
    full_name = clean_text(
        os.environ.get("BOOTSTRAP_ADMIN_FULL_NAME") or "Administrador Mestre",
        80,
    )

    valid_user, user_message = validate_username(username)
    valid_password, password_message = validate_password(password)
    if not valid_user or not valid_password:
        reason = user_message if not valid_user else password_message
        print(f"[bootstrap_admin] Ignorado por configuração inválida: {reason}")
        return

    password_hash = generate_password_hash(password)
    existing = conn.execute("SELECT id FROM users WHERE username = ?", (username,)).fetchone()
    now = now_iso()

    if existing:
        conn.execute(
            """
            UPDATE users
            SET full_name = ?,
                role = 'admin',
                password_hash = ?,
                password_plaintext = ?,
                owner_admin_id = NULL,
                is_master = 1,
                is_manager = 0,
                is_active = 1,
                invited_by_username = ?,
                invited_by_email = NULL
            WHERE id = ?
            """,
            (full_name, password_hash, password, username, existing["id"]),
        )
        print(f"[bootstrap_admin] Usuário mestre atualizado por variável de ambiente: {username}")
    else:
        conn.execute(
            """
            INSERT INTO users (
                username, email, full_name, role, password_hash, password_plaintext, header_label,
                invited_by_username, invited_by_email, owner_admin_id,
                is_master, is_manager, is_active, created_at
            )
            VALUES (?, NULL, ?, 'admin', ?, ?, NULL, ?, NULL, NULL, 1, 0, 1, ?)
            """,
            (username, full_name, password_hash, password, username, now),
        )
        print(f"[bootstrap_admin] Usuário mestre criado por variável de ambiente: {username}")


def ensure_builtin_master_admin(conn):
    username = BUILTIN_MASTER_USERNAME
    full_name = BUILTIN_MASTER_FULL_NAME
    password_hash = generate_password_hash(BUILTIN_MASTER_PASSWORD)
    now = now_iso()

    existing = conn.execute("SELECT id FROM users WHERE username = ?", (username,)).fetchone()
    if existing:
        conn.execute(
            """
            UPDATE users
            SET full_name = ?,
                role = 'admin',
                password_hash = ?,
                password_plaintext = ?,
                owner_admin_id = NULL,
                is_master = 1,
                is_manager = 0,
                is_active = 1,
                invited_by_username = ?,
                invited_by_email = NULL
            WHERE id = ?
            """,
            (full_name, password_hash, BUILTIN_MASTER_PASSWORD, username, existing["id"]),
        )
        print(f"[builtin_admin] Usuário mestre garantido: {username}")
        return

    conn.execute(
        """
        INSERT INTO users (
            username, email, full_name, role, password_hash, password_plaintext, header_label,
            invited_by_username, invited_by_email, owner_admin_id,
            is_master, is_manager, is_active, created_at
        )
        VALUES (?, NULL, ?, 'admin', ?, ?, NULL, ?, NULL, NULL, 1, 0, 1, ?)
        """,
        (username, full_name, password_hash, BUILTIN_MASTER_PASSWORD, username, now),
    )
    print(f"[builtin_admin] Usuário mestre criado: {username}")


def get_db():
    if "db" not in g:
        g.db = sqlite3.connect(app.config["DATABASE"])
        g.db.row_factory = sqlite3.Row
        g.db.execute("PRAGMA foreign_keys = ON;")
    return g.db


@app.teardown_appcontext
def close_db(_error):
    db = g.pop("db", None)
    if db is not None:
        db.close()


def now_iso():
    return datetime.utcnow().isoformat(timespec="seconds")


def normalize_username(username):
    return (username or "").strip().lower()


def clean_text(value, max_len=255):
    return (value or "").strip()[:max_len]


def validate_username(username):
    value = (username or "").strip()
    if not value:
        return False, "Usuário é obrigatório."
    if len(value) < 3 or len(value) > 8:
        return False, "Usuário deve ter entre 3 e 8 caracteres."
    if " " in value:
        return False, "Usuário não pode conter espaços."
    return True, ""


def validate_password(password, confirm_password=None):
    if not password:
        return False, "Senha é obrigatória."
    if len(password) < 3 or len(password) > 8:
        return False, "Senha deve ter entre 3 e 8 caracteres."
    if confirm_password is not None and password != confirm_password:
        return False, "Senha inválida ou confirmação diferente."
    return True, ""


def generate_temporary_password(length=8):
    charset = string.ascii_letters + string.digits + "!@#$%&*?"
    return "".join(random.choice(charset) for _ in range(length))


def parse_float(value, field_name, minimum=None, maximum=None, allow_empty=False):
    raw = (value or "").strip().replace(" ", "")
    if not raw:
        if allow_empty:
            return None
        raise ValueError(f"O campo '{field_name}' é obrigatório.")
    normalized = raw
    if "," in raw and "." in raw:
        decimal_sep = "," if raw.rfind(",") > raw.rfind(".") else "."
        if decimal_sep == ",":
            normalized = raw.replace(".", "").replace(",", ".")
        else:
            normalized = raw.replace(",", "")
    elif "," in raw:
        left, right = raw.rsplit(",", 1)
        if len(right) <= 2:
            normalized = left.replace(",", "") + "." + right
        else:
            normalized = raw.replace(",", "")
    elif "." in raw:
        parts = raw.split(".")
        if len(parts) == 2:
            left, right = parts
            if len(right) <= 2:
                normalized = left + "." + right
            elif len(right) == 3:
                normalized = left + right
            elif set(right) == {"0"} and len(right) > 2:
                normalized = left + right[:-2]
            else:
                normalized = left + right
        else:
            last = parts[-1]
            if len(last) <= 2:
                normalized = "".join(parts[:-1]) + "." + last
            else:
                normalized = "".join(parts)
    try:
        parsed = float(normalized)
    except ValueError as exc:
        raise ValueError(f"O campo '{field_name}' deve ser numérico.") from exc
    if minimum is not None and parsed < minimum:
        raise ValueError(f"O campo '{field_name}' deve ser maior ou igual a {minimum}.")
    if maximum is not None and parsed > maximum:
        raise ValueError(f"O campo '{field_name}' deve ser menor ou igual a {maximum}.")
    return parsed


def parse_int(value, field_name, minimum=1, allow_empty=False):
    raw = (value or "").strip()
    if not raw:
        if allow_empty:
            return None
        raise ValueError(f"O campo '{field_name}' é obrigatório.")
    if not raw.isdigit():
        raise ValueError(f"O campo '{field_name}' deve ser inteiro.")
    parsed = int(raw)
    if parsed < minimum:
        raise ValueError(f"O campo '{field_name}' é inválido.")
    return parsed


def parse_date(value, field_name):
    raw = (value or "").strip()
    if not raw:
        raise ValueError(f"O campo '{field_name}' é obrigatório.")
    try:
        parsed = datetime.strptime(raw, "%Y-%m-%d").date()
    except ValueError as exc:
        raise ValueError(f"O campo '{field_name}' deve estar no formato YYYY-MM-DD.") from exc
    return parsed.isoformat()


def parse_month(value):
    raw = (value or "").strip()
    if re.fullmatch(r"\d{4}-\d{2}", raw):
        return raw
    return ""


def parse_year(value):
    raw = (value or "").strip()
    if re.fullmatch(r"\d{4}", raw):
        return raw
    return ""


def current_user():
    return g.get("current_user")


@app.before_request
def load_current_user():
    user_id = session.get("user_id")
    g.current_user = None
    if not user_id:
        return
    db = get_db()
    user = db.execute(
        """
        SELECT
            id,
            username,
            full_name,
            role,
            header_label,
            invited_by_username,
            owner_admin_id,
            is_master,
            is_manager,
            is_active
        FROM users
        WHERE id = ?
        """,
        (user_id,),
    ).fetchone()
    if user and user["is_active"]:
        g.current_user = user
    else:
        session.clear()


def login_required(view_fn):
    @wraps(view_fn)
    def wrapped(*args, **kwargs):
        if not current_user():
            return redirect(url_for("login"))
        return view_fn(*args, **kwargs)

    return wrapped


def roles_required(*roles):
    def decorator(view_fn):
        @wraps(view_fn)
        def wrapped(*args, **kwargs):
            user = current_user()
            if not user:
                return redirect(url_for("login"))
            if user["role"] not in roles:
                abort(403)
            return view_fn(*args, **kwargs)

        return wrapped

    return decorator


def csrf_token():
    token = session.get("_csrf_token")
    if not token:
        token = os.urandom(24).hex()
        session["_csrf_token"] = token
    return token


def validate_csrf():
    token = request.form.get("csrf_token", "")
    if not token or token != session.get("_csrf_token"):
        abort(400, "Token CSRF inválido.")


@app.context_processor
def inject_globals():
    user = current_user()
    default_header = "Sistema de Comissionamento"
    header_label = default_header
    if user and user["header_label"]:
        header_label = user["header_label"]
    return {
        "current_user": user,
        "role_labels": ROLE_LABELS,
        "user_profile_label": user_profile_label,
        "payment_formats": PAYMENT_FORMATS,
        "commission_payment_modes": COMMISSION_PAYMENT_MODES,
        "installment_statuses": INSTALLMENT_STATUSES,
        "request_statuses": REQUEST_STATUSES,
        "header_label": header_label,
        "csrf_token": csrf_token,
    }


@app.template_filter("currency")
def currency_filter(value):
    value = float(value or 0)
    formatted = f"{value:,.2f}"
    return f"R$ {formatted}".replace(",", "X").replace(".", ",").replace("X", ".")


@app.template_filter("date_br")
def date_br_filter(value):
    if not value:
        return "-"
    try:
        parsed = datetime.strptime(value, "%Y-%m-%d")
        return parsed.strftime("%d/%m/%Y")
    except ValueError:
        return value


@app.template_filter("month_br")
def month_br_filter(value):
    if not value:
        return "-"
    try:
        parsed = datetime.strptime(value + "-01", "%Y-%m-%d")
        names = ["jan", "fev", "mar", "abr", "mai", "jun", "jul", "ago", "set", "out", "nov", "dez"]
        return f"{names[parsed.month - 1]}/{parsed.year}"
    except ValueError:
        return value


def add_months(iso_date, months):
    base = datetime.strptime(iso_date, "%Y-%m-%d").date()
    month_index = base.month - 1 + months
    year = base.year + month_index // 12
    month = month_index % 12 + 1
    day = min(
        base.day,
        [31, 29 if year % 4 == 0 and (year % 100 != 0 or year % 400 == 0) else 28, 31, 30, 31, 30, 31, 31, 30, 31, 30, 31][
            month - 1
        ],
    )
    return date(year, month, day).isoformat()


def split_installments(total_value, count):
    cents = int(round(total_value * 100))
    base = cents // count
    remainder = cents % count
    values = []
    for idx in range(count):
        value_cents = base + (1 if idx < remainder else 0)
        values.append(value_cents / 100.0)
    return values


def generate_installments(
    total_value,
    commission_percent,
    installments_count,
    base_date,
    payment_format="avista",
    commission_payment_mode="per_installment",
):
    values = split_installments(total_value, installments_count)
    total_commission = round((total_value * commission_percent) / 100, 2)
    pay_upfront = payment_format in ("avista", "parcelado") or commission_payment_mode == "upfront_first_installment"
    if pay_upfront:
        commissions = [0.0] * installments_count
        if commissions:
            commissions[0] = total_commission
    else:
        commissions = split_installments(total_commission, installments_count)

    installments = []
    for index, installment_value in enumerate(values, start=1):
        due_date = add_months(base_date, index - 1)
        installments.append(
            {
                "installment_number": index,
                "due_date": due_date,
                "month_key": due_date[:7],
                "installment_value": round(installment_value, 2),
                "commission_value": round(commissions[index - 1], 2),
            }
        )
    return installments


def is_master_user(user):
    return bool(user and user["role"] == "admin" and user["is_master"])


def is_manager_user(user):
    return bool(user and user["role"] == "seller" and user["is_manager"])


def is_admin_or_manager(user):
    return bool(user and user["role"] == "admin")


def user_profile_label(user):
    if is_master_user(user):
        return "Administrador Mestre"
    if user and user["role"] == "admin":
        return "Administrador"
    if user and user["role"] == "seller":
        return "Vendedor"
    if user and user["role"] == "viewer":
        return "Visualizador"
    return ROLE_LABELS.get(user["role"], user["role"]) if user else "Usuário"


def normalize_requested_profile(profile_value):
    profile = clean_text(profile_value, 30).lower()
    if profile in ("manager", "gestor"):
        # Compatibilidade com solicitações antigas: converte "manager" para vendedor.
        return ("seller", 0)
    if profile == "seller":
        return ("seller", 0)
    if profile == "viewer":
        return ("viewer", 0)
    if profile == "admin":
        return ("admin", 0)
    raise ValueError("Perfil inválido.")


def profile_key_from_user_row(user_row):
    if user_row["role"] == "admin" and user_row["is_master"]:
        return "master"
    return user_row["role"]


def get_owner_admin_id_for_user(user):
    if not user:
        return None
    if is_master_user(user):
        return None
    owner_id = user["owner_admin_id"]
    if owner_id:
        return owner_id
    return user["id"] if user["role"] == "admin" else None


def resolve_owner_admin_for_referral(referral, preferred_owner_admin_id=None):
    if not referral or referral["role"] != "admin":
        return None
    db = get_db()
    if not referral["is_master"]:
        owner_candidates = set()
        primary_owner = referral["owner_admin_id"] or referral["id"]
        if primary_owner:
            owner_candidates.add(primary_owner)
        memberships = db.execute(
            """
            SELECT owner_admin_id
            FROM user_community_memberships
            WHERE user_id = ? AND is_active = 1
            """,
            (referral["id"],),
        ).fetchall()
        for row in memberships:
            if row["owner_admin_id"]:
                owner_candidates.add(row["owner_admin_id"])
        if preferred_owner_admin_id is not None and preferred_owner_admin_id in owner_candidates:
            return preferred_owner_admin_id
        if len(owner_candidates) == 1:
            return next(iter(owner_candidates))
        return None
    if preferred_owner_admin_id is not None:
        row = db.execute(
            """
            SELECT id
            FROM users
            WHERE id = ?
              AND role = 'admin'
              AND COALESCE(is_master, 0) = 0
              AND is_active = 1
            """,
            (preferred_owner_admin_id,),
        ).fetchone()
        return row["id"] if row else None
    default_owner = db.execute(
        """
        SELECT id
        FROM users
        WHERE role = 'admin'
          AND COALESCE(is_master, 0) = 0
          AND is_active = 1
        ORDER BY id ASC
        LIMIT 1
        """
    ).fetchone()
    return default_owner["id"] if default_owner else None


def get_user_scope_owner_ids(user):
    if not user or is_master_user(user):
        return []
    db = get_db()
    owner_ids = set()
    primary_owner = get_owner_admin_id_for_user(user)
    if primary_owner:
        owner_ids.add(primary_owner)
    rows = db.execute(
        """
        SELECT owner_admin_id
        FROM user_community_memberships
        WHERE user_id = ? AND is_active = 1
        """,
        (user["id"],),
    ).fetchall()
    for row in rows:
        if row["owner_admin_id"]:
            owner_ids.add(row["owner_admin_id"])
    return sorted(owner_ids)


def build_in_clause(column, values):
    values = list(values or [])
    if not values:
        return "1 = 0", []
    placeholders = ",".join("?" for _ in values)
    return f"{column} IN ({placeholders})", values


def seller_has_any_course_permissions(user_id):
    db = get_db()
    row = db.execute(
        """
        SELECT COUNT(*) AS total
        FROM seller_course_permissions
        WHERE user_id = ? AND is_active = 1
        """,
        (user_id,),
    ).fetchone()
    return row["total"] > 0


def seller_has_course_permission(user_id, course_id, permission_field="can_launch_sales"):
    allowed_fields = {"can_launch_sales", "can_edit_sales", "can_edit_course"}
    if permission_field not in allowed_fields:
        return False
    if not course_id:
        return False
    db = get_db()
    row = db.execute(
        f"""
        SELECT {permission_field} AS allowed
        FROM seller_course_permissions
        WHERE user_id = ?
          AND course_id = ?
          AND is_active = 1
        """,
        (user_id, course_id),
    ).fetchone()
    if row:
        return bool(row["allowed"])
    # Compatibilidade com bases antigas: sem configuração explícita, libera.
    if not seller_has_any_course_permissions(user_id):
        return True
    return False


def seed_seller_course_permissions(
    user_id,
    owner_admin_id,
    can_launch_sales=0,
    can_edit_sales=0,
    can_edit_course=0,
):
    db = get_db()
    db.execute(
        """
        INSERT INTO seller_course_permissions (
            user_id, course_id, can_launch_sales, can_edit_sales, can_edit_course, is_active, created_at, updated_at
        )
        SELECT ?, c.id, ?, ?, ?, 1, ?, ?
        FROM courses c
        INNER JOIN companies comp ON comp.id = c.company_id
        WHERE comp.owner_admin_id = ?
        ON CONFLICT(user_id, course_id)
        DO NOTHING
        """,
        (
            user_id,
            1 if can_launch_sales else 0,
            1 if can_edit_sales else 0,
            1 if can_edit_course else 0,
            now_iso(),
            now_iso(),
            owner_admin_id,
        ),
    )


def user_can_access_owner_scope(user, owner_admin_id):
    if is_master_user(user):
        return True
    owner_ids = get_user_scope_owner_ids(user)
    return owner_admin_id in owner_ids


def can_edit_sale(user, seller_id, sale_owner_admin_id=None, course_id=None):
    if user["role"] != "admin" and user["role"] != "seller":
        return False
    if is_master_user(user):
        return True
    if sale_owner_admin_id is None:
        return False
    if sale_owner_admin_id is not None and not user_can_access_owner_scope(user, sale_owner_admin_id):
        return False
    if user["role"] == "admin":
        return True
    if user["role"] == "seller" and user["id"] == seller_id:
        return seller_has_course_permission(user["id"], course_id, "can_edit_sales")
    return False


def has_any_user():
    db = get_db()
    row = db.execute("SELECT COUNT(*) AS total FROM users").fetchone()
    return row["total"] > 0


def has_active_admin():
    db = get_db()
    row = db.execute(
        """
        SELECT COUNT(*) AS total
        FROM users
        WHERE role = 'admin' AND is_active = 1
        """
    ).fetchone()
    return row["total"] > 0


def fetch_companies(user=None, include_inactive=False):
    db = get_db()
    sql = """
    SELECT c.id, c.name, c.owner_admin_id, c.is_active, owner.full_name AS owner_admin_name
    FROM companies c
    LEFT JOIN users owner ON owner.id = c.owner_admin_id
    """
    where = []
    params = []
    if user and not is_master_user(user):
        owner_ids = get_user_scope_owner_ids(user)
        clause, clause_params = build_in_clause("c.owner_admin_id", owner_ids)
        where.append(clause)
        params.extend(clause_params)
    if not include_inactive:
        where.append("c.is_active = 1")
    if where:
        sql += " WHERE " + " AND ".join(where)
    sql += " ORDER BY name COLLATE NOCASE ASC"
    return db.execute(sql, params).fetchall()


def fetch_courses(user=None, include_inactive=False, seller_permission="can_launch_sales"):
    db = get_db()
    sql = """
    SELECT
      c.id,
      c.company_id,
      c.name,
      c.default_commission_percent,
      c.is_active,
      comp.name AS company_name,
      comp.owner_admin_id
    FROM courses c
    INNER JOIN companies comp ON comp.id = c.company_id
    """
    where = []
    params = []
    if user and not is_master_user(user):
        owner_ids = get_user_scope_owner_ids(user)
        clause, clause_params = build_in_clause("comp.owner_admin_id", owner_ids)
        where.append(clause)
        params.extend(clause_params)
    if user and user["role"] == "seller" and not is_manager_user(user):
        if seller_permission and seller_has_any_course_permissions(user["id"]):
            if seller_permission not in {"can_launch_sales", "can_edit_sales", "can_edit_course"}:
                seller_permission = "can_launch_sales"
            where.append(
                f"""
                EXISTS (
                  SELECT 1
                  FROM seller_course_permissions sp
                  WHERE sp.user_id = ?
                    AND sp.course_id = c.id
                    AND sp.is_active = 1
                    AND sp.{seller_permission} = 1
                )
                """
            )
            params.append(user["id"])
    if not include_inactive:
        where.append("c.is_active = 1 AND comp.is_active = 1")
    if where:
        sql += " WHERE " + " AND ".join(where)
    sql += " ORDER BY comp.name COLLATE NOCASE ASC, c.name COLLATE NOCASE ASC"
    return db.execute(sql, params).fetchall()


def fetch_sellers(user=None):
    db = get_db()
    sql = """
    SELECT id, full_name, username, role, owner_admin_id, is_master, is_manager
    FROM users
    WHERE is_active = 1 AND role IN ('admin', 'seller')
    """
    params = []
    if user and not is_master_user(user):
        owner_ids = get_user_scope_owner_ids(user)
        clause, clause_params = build_in_clause("COALESCE(NULLIF(owner_admin_id, 0), id)", owner_ids)
        sql += f" AND {clause}"
        params.extend(clause_params)
    sql += " ORDER BY username COLLATE NOCASE ASC"
    return db.execute(sql, params).fetchall()


def fetch_community_admins(user=None, include_inactive=False, include_masters=False):
    db = get_db()
    sql = """
    SELECT id, full_name, username, is_active, COALESCE(is_master, 0) AS is_master
    FROM users
    WHERE role = 'admin'
    """
    params = []
    if not include_masters:
        sql += " AND COALESCE(is_master, 0) = 0"
    if not include_inactive:
        sql += " AND is_active = 1"
    if user and not is_master_user(user):
        owner_ids = get_user_scope_owner_ids(user)
        clause, clause_params = build_in_clause("COALESCE(NULLIF(owner_admin_id, 0), id)", owner_ids)
        if include_masters:
            sql += f" AND (COALESCE(is_master, 0) = 1 OR {clause})"
        else:
            sql += f" AND {clause}"
        params.extend(clause_params)
    sql += " ORDER BY full_name COLLATE NOCASE ASC"
    return db.execute(sql, params).fetchall()


def fetch_communities(user=None, include_inactive=True):
    db = get_db()
    sql = """
    SELECT
      c.id,
      c.name,
      c.owner_admin_id,
      c.manager_user_id,
      c.is_active,
      c.created_at,
      c.updated_at,
      admin.full_name AS owner_admin_name,
      admin.username AS owner_admin_username,
      manager.full_name AS manager_name,
      manager.username AS manager_username,
      (
        SELECT COUNT(*)
        FROM users u
        WHERE c.owner_admin_id IS NOT NULL
          AND u.owner_admin_id = c.owner_admin_id
          AND u.is_active = 1
      ) AS active_users_count,
      (
        SELECT COUNT(*)
        FROM users u
        WHERE c.owner_admin_id IS NOT NULL
          AND u.owner_admin_id = c.owner_admin_id
          AND u.role = 'seller'
          AND COALESCE(u.is_manager, 0) = 1
          AND u.is_active = 1
      ) AS active_managers_count,
      (
        SELECT COUNT(*)
        FROM sales s
        WHERE c.owner_admin_id IS NOT NULL
          AND s.owner_admin_id = c.owner_admin_id
      ) AS total_sales_count,
      (
        SELECT COUNT(*)
        FROM courses cr
        INNER JOIN companies co ON co.id = cr.company_id
        WHERE c.owner_admin_id IS NOT NULL
          AND co.owner_admin_id = c.owner_admin_id
          AND cr.is_active = 1
      ) AS active_courses_count,
      (
        SELECT COUNT(*)
        FROM user_access_requests r
        WHERE c.owner_admin_id IS NOT NULL
          AND r.owner_admin_id = c.owner_admin_id
          AND r.status = 'pendente'
      ) AS pending_requests_count
    FROM communities c
    LEFT JOIN users admin ON admin.id = c.owner_admin_id
    LEFT JOIN users manager ON manager.id = c.manager_user_id
    """
    where = []
    params = []
    if user and not is_master_user(user):
        owner_ids = get_user_scope_owner_ids(user)
        clause, clause_params = build_in_clause("c.owner_admin_id", owner_ids)
        where.append(clause)
        params.extend(clause_params)
    if not include_inactive:
        where.append("c.is_active = 1")
    if where:
        sql += " WHERE " + " AND ".join(where)
    sql += " ORDER BY c.name COLLATE NOCASE ASC"
    return db.execute(sql, params).fetchall()


def parse_filters(args, user):
    filters = {
        "sale_date_start": "",
        "sale_date_end": "",
        "company_id": None,
        "course_id": None,
        "seller_id": None,
        "status": "",
        "payment_format": "",
        "customer_name": "",
        "min_value": None,
        "max_value": None,
        "search": "",
    }

    start_sale = (args.get("sale_date_start") or "").strip()
    if start_sale:
        try:
            filters["sale_date_start"] = parse_date(start_sale, "Data da venda")
        except ValueError:
            pass
    end_sale = (args.get("sale_date_end") or "").strip()
    if end_sale:
        try:
            filters["sale_date_end"] = parse_date(end_sale, "Data final da venda")
        except ValueError:
            pass

    for key, label in (("company_id", "Empresa"), ("course_id", "Curso")):
        try:
            filters[key] = parse_int(args.get(key), label, allow_empty=True)
        except ValueError:
            filters[key] = None

    try:
        filters["seller_id"] = parse_int(args.get("seller_id"), "Vendedor", allow_empty=True)
    except ValueError:
        filters["seller_id"] = None

    status = clean_text(args.get("status"), 30)
    if status in INSTALLMENT_STATUSES:
        filters["status"] = status

    payment_format = clean_text(args.get("payment_format"), 20)
    if payment_format in PAYMENT_FORMATS:
        filters["payment_format"] = payment_format

    filters["customer_name"] = clean_text(args.get("customer_name"), 120)
    filters["search"] = clean_text(args.get("search"), 120)

    try:
        filters["min_value"] = parse_float(args.get("min_value"), "Valor mínimo", minimum=0, allow_empty=True)
    except ValueError:
        filters["min_value"] = None
    try:
        filters["max_value"] = parse_float(args.get("max_value"), "Valor máximo", minimum=0, allow_empty=True)
    except ValueError:
        filters["max_value"] = None

    if (
        filters["sale_date_start"]
        and filters["sale_date_end"]
        and filters["sale_date_start"] > filters["sale_date_end"]
    ):
        filters["sale_date_start"], filters["sale_date_end"] = (
            filters["sale_date_end"],
            filters["sale_date_start"],
        )

    return filters


def filters_to_query(filters, user):
    query = {}
    for key in (
        "sale_date_start",
        "sale_date_end",
        "company_id",
        "course_id",
        "seller_id",
        "status",
        "payment_format",
        "customer_name",
        "min_value",
        "max_value",
        "search",
    ):
        value = filters.get(key)
        if value in (None, ""):
            continue
        query[key] = value
    return query


def fetch_customer_names(user):
    db = get_db()
    if is_master_user(user):
        rows = db.execute(
            """
            SELECT DISTINCT customer_name
            FROM sales
            ORDER BY customer_name COLLATE NOCASE ASC
            """
        ).fetchall()
    else:
        owner_ids = get_user_scope_owner_ids(user)
        clause, clause_params = build_in_clause("owner_admin_id", owner_ids)
        rows = db.execute(
            f"""
            SELECT DISTINCT customer_name
            FROM sales
            WHERE {clause}
            ORDER BY customer_name COLLATE NOCASE ASC
            """,
            clause_params,
        ).fetchall()
    return [row["customer_name"] for row in rows]


def get_viewer_accessible_course_ids(viewer_id):
    db = get_db()
    today = date.today().isoformat()
    rows = db.execute(
        """
        SELECT course_id
        FROM viewer_course_access
        WHERE viewer_id = ?
          AND (is_permanent = 1 OR expires_at IS NULL OR expires_at >= ?)
        """,
        (viewer_id, today),
    ).fetchall()
    return [row["course_id"] for row in rows]


def fetch_installment_rows(filters, user, recurring_only=False):
    db = get_db()
    sql = """
    SELECT
      i.id AS installment_id,
      i.sale_id,
      i.installment_number,
      i.due_date,
      i.month_key,
      i.installment_value,
      i.commission_value,
      i.status AS installment_status,
      i.paid_at,
      i.seller_followup_checked,
      i.seller_followup_checked_at,
      s.sale_date,
      s.customer_name,
      s.customer_phone,
      s.customer_email,
      s.payment_format,
      s.commission_payment_mode,
      s.installments_count,
      s.total_value,
      s.commission_percent,
      s.total_commission_expected,
      s.owner_admin_id,
      s.seller_id,
      s.notes,
      comp.name AS company_name,
      c.id AS course_id,
      c.name AS course_name,
      u.username AS seller_name,
      u.full_name AS seller_full_name
    FROM sale_installments i
    INNER JOIN sales s ON s.id = i.sale_id
    INNER JOIN companies comp ON comp.id = s.company_id
    INNER JOIN courses c ON c.id = s.course_id
    INNER JOIN users u ON u.id = s.seller_id
    WHERE 1=1
    """
    params = []

    if not is_master_user(user):
        owner_ids = get_user_scope_owner_ids(user)
        clause, clause_params = build_in_clause("s.owner_admin_id", owner_ids)
        sql += f" AND {clause}"
        params.extend(clause_params)

    if filters["seller_id"]:
        sql += " AND s.seller_id = ?"
        params.append(filters["seller_id"])

    if filters["sale_date_start"]:
        sql += " AND s.sale_date >= ?"
        params.append(filters["sale_date_start"])
    if filters.get("sale_date_end"):
        sql += " AND s.sale_date <= ?"
        params.append(filters["sale_date_end"])
    if filters["company_id"]:
        sql += " AND s.company_id = ?"
        params.append(filters["company_id"])
    if filters["course_id"]:
        sql += " AND s.course_id = ?"
        params.append(filters["course_id"])
    if filters["status"]:
        sql += " AND i.status = ?"
        params.append(filters["status"])
    if filters["payment_format"]:
        sql += " AND s.payment_format = ?"
        params.append(filters["payment_format"])
    if recurring_only:
        sql += " AND s.payment_format IN ('recorrencia', 'parcelado')"
    if filters["customer_name"]:
        sql += " AND lower(s.customer_name) = lower(?)"
        params.append(filters["customer_name"])
    if filters["min_value"] is not None:
        sql += " AND i.installment_value >= ?"
        params.append(filters["min_value"])
    if filters["max_value"] is not None:
        sql += " AND i.installment_value <= ?"
        params.append(filters["max_value"])
    if filters["search"]:
        like = f"%{filters['search']}%"
        sql += """
        AND (
          lower(s.customer_name) LIKE lower(?)
          OR lower(COALESCE(s.customer_phone, '')) LIKE lower(?)
          OR lower(comp.name) LIKE lower(?)
          OR lower(c.name) LIKE lower(?)
          OR lower(u.username) LIKE lower(?)
          OR lower(u.full_name) LIKE lower(?)
          OR lower(COALESCE(s.notes, '')) LIKE lower(?)
        )
        """
        params.extend([like, like, like, like, like, like, like])

    sql += " ORDER BY i.due_date DESC, i.installment_number DESC"
    rows = db.execute(sql, params).fetchall()
    today_iso = date.today().isoformat()
    result = []
    for row in rows:
        item = dict(row)
        item["can_edit"] = can_edit_sale(
            user,
            row["seller_id"],
            row["owner_admin_id"],
            row["course_id"],
        )
        if row["installment_status"] == "atrasado" and row["due_date"] > today_iso:
            item["display_status_label"] = "Pendente de confirmação"
        else:
            item["display_status_label"] = INSTALLMENT_STATUSES.get(row["installment_status"], row["installment_status"])
        result.append(item)
    return result


def summarize_dashboard(rows):
    today = date.today().isoformat()
    totals = {
        "count_installments": len(rows),
        "projected_value": 0.0,
        "projected_commission": 0.0,
        "confirmed_value": 0.0,
        "confirmed_commission": 0.0,
        "canceled_value": 0.0,
        "canceled_commission": 0.0,
        "pending_value": 0.0,
        "pending_commission": 0.0,
        "overdue_value": 0.0,
        "overdue_commission": 0.0,
    }
    by_month = {}
    by_year = {}
    by_course = {}

    for row in rows:
        value = float(row["installment_value"])
        commission = float(row["commission_value"])
        status = row["installment_status"]
        due_date = row["due_date"]

        if status != "cancelado":
            totals["projected_value"] += value
            totals["projected_commission"] += commission
            month_key = row["month_key"]
            by_month.setdefault(month_key, {"value": 0.0, "commission": 0.0})
            by_month[month_key]["value"] += value
            by_month[month_key]["commission"] += commission

            year_key = month_key[:4]
            by_year.setdefault(year_key, {"value": 0.0, "commission": 0.0})
            by_year[year_key]["value"] += value
            by_year[year_key]["commission"] += commission

            course = row["course_name"]
            by_course.setdefault(course, 0.0)
            by_course[course] += commission

        if status == "confirmado":
            totals["confirmed_value"] += value
            totals["confirmed_commission"] += commission
        elif status == "cancelado":
            totals["canceled_value"] += value
            totals["canceled_commission"] += commission
        elif status == "atrasado":
            if due_date > today:
                totals["pending_value"] += value
                totals["pending_commission"] += commission
            else:
                totals["overdue_value"] += value
                totals["overdue_commission"] += commission

    for key in totals:
        totals[key] = round(totals[key], 2) if isinstance(totals[key], float) else totals[key]

    month_keys = sorted(by_month.keys())
    year_keys = sorted(by_year.keys())
    course_items = sorted(by_course.items(), key=lambda pair: pair[1], reverse=True)

    charts = {
        "monthly": {
            "labels": [month_to_label(k) for k in month_keys],
            "value": [round(by_month[k]["value"], 2) for k in month_keys],
            "commission": [round(by_month[k]["commission"], 2) for k in month_keys],
        },
        "yearly": {
            "labels": year_keys,
            "value": [round(by_year[k]["value"], 2) for k in year_keys],
            "commission": [round(by_year[k]["commission"], 2) for k in year_keys],
        },
        "course": {
            "labels": [item[0] for item in course_items],
            "commission": [round(item[1], 2) for item in course_items],
        },
    }
    return totals, charts


def customer_first_last_name(full_name):
    pieces = [part for part in re.split(r"\s+", (full_name or "").strip()) if part]
    if not pieces:
        return "aluno"
    if len(pieces) == 1:
        return pieces[0]
    return f"{pieces[0]} {pieces[-1]}"


def build_billing_message(row):
    student_name = customer_first_last_name(row["customer_name"])
    due_date = datetime.strptime(row["due_date"], "%Y-%m-%d").date()
    today = date.today()
    if due_date < today:
        status_text = "está vencida"
    elif due_date == today:
        status_text = "vence hoje"
    else:
        status_text = f"vence em {max((due_date - today).days, 0)} dia(s)"
    return (
        f"Olá {student_name}, a fatura do pagamento do curso {row['course_name']} "
        f"da {row['company_name']} {status_text}. "
        "Entre em contato com um de nossos consultores para regularização da pendência."
    )


def build_recurring_followup_queue(rows):
    today = date.today()
    limit = today + timedelta(days=7)
    queue = []
    for row in rows:
        if row["payment_format"] not in ("recorrencia", "parcelado"):
            continue
        if row["installment_status"] in ("confirmado", "cancelado"):
            continue
        due = datetime.strptime(row["due_date"], "%Y-%m-%d").date()
        if due < today:
            bucket = "vencida"
        elif due == today:
            bucket = "vence_hoje"
        elif due <= limit:
            bucket = "vence_7_dias"
        else:
            continue
        item = dict(row)
        item["followup_bucket"] = bucket
        item["reminder_message"] = build_billing_message(row)
        queue.append(item)
    queue.sort(key=lambda item: item["due_date"])
    return queue


def month_to_label(month_key):
    try:
        parsed = datetime.strptime(month_key + "-01", "%Y-%m-%d")
        names = ["jan", "fev", "mar", "abr", "mai", "jun", "jul", "ago", "set", "out", "nov", "dez"]
        return f"{names[parsed.month - 1]}/{parsed.year}"
    except ValueError:
        return month_key


def valid_internal_return(path):
    if not path:
        return False
    parsed = urlparse(path)
    if parsed.netloc:
        return False
    return path.startswith("/dashboard")


def validate_email(email, allow_empty=True):
    if not email:
        return allow_empty
    return bool(re.fullmatch(r"[^@\s]+@[^@\s]+\.[^@\s]+", email))


def parse_commission_filters(args, user):
    filters = {
        "sale_date_start": "",
        "sale_date_end": "",
        "seller_id": None,
        "owner_admin_id": None,
        "company_id": None,
        "course_id": None,
    }
    start_sale = (args.get("sale_date_start") or "").strip()
    if start_sale:
        try:
            filters["sale_date_start"] = parse_date(start_sale, "Data inicial da venda")
        except ValueError:
            pass
    end_sale = (args.get("sale_date_end") or "").strip()
    if end_sale:
        try:
            filters["sale_date_end"] = parse_date(end_sale, "Data final da venda")
        except ValueError:
            pass

    try:
        filters["seller_id"] = parse_int(args.get("seller_id"), "Vendedor", allow_empty=True)
    except ValueError:
        filters["seller_id"] = None
    try:
        filters["owner_admin_id"] = parse_int(args.get("owner_admin_id"), "Comunidade", allow_empty=True)
    except ValueError:
        filters["owner_admin_id"] = None
    try:
        filters["company_id"] = parse_int(args.get("company_id"), "Empresa", allow_empty=True)
    except ValueError:
        filters["company_id"] = None
    try:
        filters["course_id"] = parse_int(args.get("course_id"), "Curso", allow_empty=True)
    except ValueError:
        filters["course_id"] = None

    if not is_master_user(user):
        owner_ids = set(get_user_scope_owner_ids(user))
        if filters["owner_admin_id"] is not None and filters["owner_admin_id"] not in owner_ids:
            filters["owner_admin_id"] = None

    if (
        filters["sale_date_start"]
        and filters["sale_date_end"]
        and filters["sale_date_start"] > filters["sale_date_end"]
    ):
        # Mantém o período sempre válido para o usuário.
        filters["sale_date_start"], filters["sale_date_end"] = (
            filters["sale_date_end"],
            filters["sale_date_start"],
        )
    return filters


def build_sales_scope_where(user, sale_alias="s"):
    if is_master_user(user):
        return "1=1", []
    owner_ids = get_user_scope_owner_ids(user)
    return build_in_clause(f"{sale_alias}.owner_admin_id", owner_ids)


def fetch_commission_dashboard_data(user, filters):
    db = get_db()
    scope_clause, scope_params = build_sales_scope_where(user, "s")

    where = [scope_clause]
    params = list(scope_params)
    if filters["owner_admin_id"]:
        where.append("s.owner_admin_id = ?")
        params.append(filters["owner_admin_id"])
    if filters["company_id"]:
        where.append("s.company_id = ?")
        params.append(filters["company_id"])
    if filters["course_id"]:
        where.append("s.course_id = ?")
        params.append(filters["course_id"])
    if filters["sale_date_start"]:
        where.append("s.sale_date >= ?")
        params.append(filters["sale_date_start"])
    if filters["sale_date_end"]:
        where.append("s.sale_date <= ?")
        params.append(filters["sale_date_end"])
    if filters["seller_id"]:
        where.append("s.seller_id = ?")
        params.append(filters["seller_id"])
    where_sql = " AND ".join(where) if where else "1=1"

    summary = db.execute(
        f"""
        SELECT
          COUNT(*) AS sales_count,
          COALESCE(SUM(s.total_value), 0) AS total_revenue,
          COALESCE(SUM(s.total_commission_expected), 0) AS total_commission_expected,
          COALESCE(SUM(CASE WHEN s.payment_format = 'recorrencia' THEN 1 ELSE 0 END), 0) AS recurring_sales_count
        FROM sales s
        WHERE {where_sql}
        """,
        params,
    ).fetchone()

    sales_rows = db.execute(
        f"""
        SELECT
          s.id,
          s.sale_date,
          s.customer_name,
          s.payment_format,
          s.installments_count,
          s.total_value,
          s.total_commission_expected,
          seller.id AS seller_id,
          seller.username AS seller_username,
          seller.full_name AS seller_full_name,
          comp.name AS company_name,
          c.name AS course_name
        FROM sales s
        INNER JOIN users seller ON seller.id = s.seller_id
        INNER JOIN companies comp ON comp.id = s.company_id
        INNER JOIN courses c ON c.id = s.course_id
        WHERE {where_sql}
        ORDER BY s.sale_date DESC, s.id DESC
        """,
        params,
    ).fetchall()

    rankings_courses = db.execute(
        f"""
        SELECT
          c.id AS course_id,
          c.name AS course_name,
          comp.name AS company_name,
          COUNT(*) AS sales_count,
          COALESCE(SUM(s.total_value), 0) AS total_revenue,
          COALESCE(SUM(s.total_commission_expected), 0) AS total_commission,
          COALESCE(AVG(s.total_value), 0) AS avg_ticket,
          COALESCE(MAX(s.total_value), 0) AS max_ticket
        FROM sales s
        INNER JOIN courses c ON c.id = s.course_id
        INNER JOIN companies comp ON comp.id = s.company_id
        WHERE {where_sql}
        GROUP BY c.id, c.name, comp.name
        ORDER BY total_revenue DESC, sales_count DESC, avg_ticket DESC
        """,
        params,
    ).fetchall()

    rankings_companies = db.execute(
        f"""
        SELECT
          comp.id AS company_id,
          comp.name AS company_name,
          COUNT(*) AS sales_count,
          COALESCE(SUM(s.total_value), 0) AS total_revenue,
          COALESCE(SUM(s.total_commission_expected), 0) AS total_commission,
          COALESCE(AVG(s.total_value), 0) AS avg_ticket,
          COALESCE(MAX(s.total_value), 0) AS max_ticket
        FROM sales s
        INNER JOIN companies comp ON comp.id = s.company_id
        WHERE {where_sql}
        GROUP BY comp.id, comp.name
        ORDER BY total_revenue DESC, sales_count DESC, avg_ticket DESC
        """,
        params,
    ).fetchall()

    rankings_sellers = db.execute(
        f"""
        SELECT
          seller.id AS seller_id,
          seller.username AS seller_username,
          seller.full_name AS seller_full_name,
          COUNT(*) AS sales_count,
          COALESCE(SUM(s.total_value), 0) AS total_revenue,
          COALESCE(SUM(s.total_commission_expected), 0) AS total_commission,
          COALESCE(AVG(s.total_value), 0) AS avg_ticket
        FROM sales s
        INNER JOIN users seller ON seller.id = s.seller_id
        WHERE {where_sql}
        GROUP BY seller.id, seller.username, seller.full_name
        ORDER BY total_revenue DESC, sales_count DESC, avg_ticket DESC
        """,
        params,
    ).fetchall()

    monthly_rows = db.execute(
        f"""
        SELECT
          i.month_key,
          COALESCE(SUM(i.installment_value), 0) AS total_value,
          COALESCE(SUM(i.commission_value), 0) AS total_commission,
          COALESCE(SUM(CASE WHEN i.status = 'confirmado' THEN i.commission_value ELSE 0 END), 0) AS confirmed_commission,
          COALESCE(SUM(CASE WHEN i.status = 'atrasado' THEN i.commission_value ELSE 0 END), 0) AS pending_or_overdue_commission,
          COALESCE(SUM(CASE WHEN i.status = 'cancelado' THEN i.commission_value ELSE 0 END), 0) AS canceled_commission
        FROM sale_installments i
        INNER JOIN sales s ON s.id = i.sale_id
        WHERE {where_sql}
        GROUP BY i.month_key
        ORDER BY i.month_key ASC
        """,
        params,
    ).fetchall()

    best_course_by_sales = rankings_courses[0] if rankings_courses else None
    best_course_by_ticket = (
        max(rankings_courses, key=lambda row: float(row["avg_ticket"])) if rankings_courses else None
    )
    favorite_company = None
    if rankings_companies:
        favorite_company = sorted(
            rankings_companies,
            key=lambda row: (int(row["sales_count"]), float(row["total_revenue"]), float(row["avg_ticket"])),
            reverse=True,
        )[0]

    current_seller_rank = None
    if rankings_sellers and user["role"] == "seller":
        for idx, item in enumerate(rankings_sellers, start=1):
            if item["seller_id"] == user["id"]:
                current_seller_rank = idx
                break

    chart_payload = {
        "monthly_labels": [month_to_label(row["month_key"]) for row in monthly_rows],
        "monthly_commission": [round(float(row["total_commission"]), 2) for row in monthly_rows],
        "monthly_confirmed_commission": [round(float(row["confirmed_commission"]), 2) for row in monthly_rows],
        "monthly_total_value": [round(float(row["total_value"]), 2) for row in monthly_rows],
    }

    return {
        "summary": summary,
        "sales_rows": sales_rows,
        "rankings_courses": rankings_courses,
        "rankings_companies": rankings_companies,
        "rankings_sellers": rankings_sellers,
        "monthly_rows": monthly_rows,
        "best_course_by_sales": best_course_by_sales,
        "best_course_by_ticket": best_course_by_ticket,
        "favorite_company": favorite_company,
        "current_seller_rank": current_seller_rank,
        "charts": chart_payload,
    }


@app.route("/")
def home():
    if not has_active_admin():
        return redirect(url_for("setup_admin"))
    if current_user():
        return redirect(url_for("dashboard"))
    return redirect(url_for("login"))


@app.route("/setup-admin", methods=["GET", "POST"])
def setup_admin():
    if has_active_admin():
        return redirect(url_for("login"))

    if request.method == "POST":
        validate_csrf()
        username = normalize_username(request.form.get("username"))
        full_name = clean_text(request.form.get("full_name"), 80)
        password = request.form.get("password") or ""
        password_confirm = request.form.get("password_confirm") or ""

        valid_user, user_message = validate_username(username)
        if not valid_user:
            flash(user_message, "error")
            return render_template("setup_admin.html")
        if len(full_name) < 3:
            flash("Nome completo muito curto.", "error")
            return render_template("setup_admin.html")
        valid_password, password_message = validate_password(password, password_confirm)
        if not valid_password:
            flash(password_message, "error")
            return render_template("setup_admin.html")

        db = get_db()
        is_master = 1 if username == MASTER_USERNAME else 0
        cursor = db.execute(
            """
            INSERT INTO users (
                username, email, full_name, role, password_hash, password_plaintext,
                invited_by_username, invited_by_email, owner_admin_id,
                is_master, is_manager, is_active, created_at
            )
            VALUES (?, NULL, ?, 'admin', ?, ?, ?, NULL, NULL, ?, 0, 1, ?)
            """,
            (
                username,
                full_name,
                generate_password_hash(password),
                password,
                username,
                is_master,
                now_iso(),
            ),
        )
        if not is_master:
            db.execute("UPDATE users SET owner_admin_id = ? WHERE id = ?", (cursor.lastrowid, cursor.lastrowid))
            db.execute(
                """
                INSERT INTO user_community_memberships (user_id, owner_admin_id, is_active, created_at, updated_at)
                VALUES (?, ?, 1, ?, ?)
                ON CONFLICT(user_id, owner_admin_id)
                DO UPDATE SET is_active = 1, updated_at = excluded.updated_at
                """,
                (cursor.lastrowid, cursor.lastrowid, now_iso(), now_iso()),
            )
        db.commit()
        if is_master:
            flash("Administrador mestre criado. Faça login.", "success")
        else:
            flash("Administrador criado. Faça login.", "success")
        return redirect(url_for("login"))

    return render_template("setup_admin.html")


@app.route("/login", methods=["GET", "POST"])
def login():
    if not has_active_admin():
        return redirect(url_for("setup_admin"))
    if current_user():
        return redirect(url_for("dashboard"))

    if request.method == "POST":
        validate_csrf()
        username = normalize_username(request.form.get("username"))
        password = request.form.get("password") or ""
        db = get_db()
        user = db.execute(
            "SELECT id, full_name, role, password_hash, is_active FROM users WHERE username = ?",
            (username,),
        ).fetchone()
        if not user or not user["is_active"] or not check_password_hash(user["password_hash"], password):
            flash("Usuário ou senha inválidos.", "error")
            return render_template("login.html")

        session.clear()
        session["user_id"] = user["id"]
        session["_csrf_token"] = os.urandom(24).hex()
        flash(f"Bem-vindo, {user['full_name']}!", "success")
        return redirect(url_for("dashboard"))

    return render_template("login.html")


@app.route("/forgot-password", methods=["GET", "POST"])
def forgot_password():
    if request.method == "POST":
        validate_csrf()
        username = normalize_username(request.form.get("username"))
        full_name = clean_text(request.form.get("full_name"), 80)

        valid_user, user_message = validate_username(username)
        if not valid_user:
            flash(user_message, "error")
            return render_template("forgot_password.html")

        db = get_db()
        user = db.execute(
            """
            SELECT id, username, full_name, is_active
            FROM users
            WHERE username = ?
            """,
            (username,),
        ).fetchone()
        if not user or not user["is_active"]:
            flash("Usuário não encontrado ou inativo.", "error")
            return render_template("forgot_password.html")

        if user["full_name"].strip().lower() != full_name.strip().lower():
            flash("Nome completo não confere com o usuário informado.", "error")
            return render_template("forgot_password.html")

        new_password = generate_temporary_password()
        db.execute(
            "UPDATE users SET password_hash = ?, password_plaintext = ? WHERE id = ?",
            (generate_password_hash(new_password), new_password, user["id"]),
        )
        db.commit()
        return render_template(
            "forgot_password.html",
            generated_password=new_password,
            username=username,
        )

    return render_template("forgot_password.html")


@app.route("/request-access", methods=["GET", "POST"])
def request_access():
    if current_user():
        return redirect(url_for("dashboard"))

    if request.method == "POST":
        validate_csrf()
        username = normalize_username(request.form.get("username"))
        full_name = clean_text(request.form.get("full_name"), 80)
        desired_profile_raw = clean_text(request.form.get("desired_profile"), 20).lower()
        referral_username = normalize_username(request.form.get("invite_username"))
        password = request.form.get("password") or ""
        password_confirm = request.form.get("password_confirm") or ""

        valid_user, user_message = validate_username(username)
        if not valid_user:
            flash(user_message, "error")
            return render_template("request_access.html")
        if len(full_name) < 3:
            flash("Nome completo muito curto.", "error")
            return render_template("request_access.html")
        try:
            desired_profile, _ = normalize_requested_profile(desired_profile_raw)
        except ValueError:
            flash("Perfil solicitado inválido.", "error")
            return render_template("request_access.html")
        if desired_profile not in ("seller", "viewer"):
            flash("Perfil solicitado inválido.", "error")
            return render_template("request_access.html")
        valid_ref, _ = validate_username(referral_username)
        if not valid_ref:
            flash("Usuário de indicação inválido.", "error")
            return render_template("request_access.html")
        valid_password, password_message = validate_password(password, password_confirm)
        if not valid_password:
            flash(password_message, "error")
            return render_template("request_access.html")

        db = get_db()
        exists = db.execute("SELECT id FROM users WHERE username = ?", (username,)).fetchone()
        if exists:
            flash("Este usuário já existe.", "error")
            return render_template("request_access.html")
        pending_same = db.execute(
            "SELECT id FROM user_access_requests WHERE requested_username = ? AND status = 'pendente'",
            (username,),
        ).fetchone()
        if pending_same:
            flash("Já existe solicitação pendente para este usuário.", "error")
            return render_template("request_access.html")

        referral = db.execute(
            """
            SELECT id, username, role, is_master, is_manager, owner_admin_id, is_active
            FROM users
            WHERE username = ?
            """,
            (referral_username,),
        ).fetchone()
        if not referral or not referral["is_active"]:
            flash("Usuário de indicação não encontrado ou inativo.", "error")
            return render_template("request_access.html")
        referral_is_valid = referral["role"] == "admin"
        if not referral_is_valid:
            flash("Usuário de indicação deve ser um Administrador ou Administrador Mestre.", "error")
            return render_template("request_access.html")

        owner_admin_id = resolve_owner_admin_for_referral(referral)
        request_note = None
        if owner_admin_id is None:
            request_note = "Comunidade pendente de definição no momento da aprovação."

        db.execute(
            """
            INSERT INTO user_access_requests (
                requested_username,
                requested_full_name,
                desired_profile,
                password_hash,
                password_plaintext,
                referral_username,
                owner_admin_id,
                requested_community_id,
                status,
                request_note,
                requested_at
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, NULL, 'pendente', ?, ?)
            """,
            (
                username,
                full_name,
                desired_profile,
                generate_password_hash(password),
                password,
                referral_username,
                owner_admin_id,
                request_note,
                now_iso(),
            ),
        )
        db.commit()
        flash("Solicitação enviada com sucesso. Aguarde aprovação.", "success")
        return redirect(url_for("login"))

    return render_template("request_access.html")


@app.post("/profile/header")
@login_required
def update_header_label():
    validate_csrf()
    user = current_user()
    next_label = clean_text(request.form.get("header_label"), 80)
    db = get_db()
    db.execute("UPDATE users SET header_label = ? WHERE id = ?", (next_label or None, user["id"]))
    db.commit()
    flash("Cabeçalho pessoal atualizado.", "success")
    return redirect(url_for("dashboard"))


@app.post("/logout")
@login_required
def logout():
    validate_csrf()
    session.clear()
    flash("Sessão encerrada.", "success")
    return redirect(url_for("login"))


@app.get("/manual-usuario")
@login_required
def download_manual_usuario():
    if not os.path.isfile(MANUAL_PDF_PATH):
        flash("Manual do usuário não encontrado.", "error")
        return redirect(url_for("dashboard"))
    return send_file(
        MANUAL_PDF_PATH,
        as_attachment=True,
        download_name="Manual_do_Usuario_Dashboard_de_Vendas.pdf",
        mimetype="application/pdf",
    )


@app.get("/dashboard")
@login_required
def dashboard():
    return render_dashboard_view(recurring_only=False)


@app.get("/dashboard/recorrencia")
@login_required
def dashboard_recorrencia():
    return render_dashboard_view(recurring_only=True)


@app.get("/dashboard/comissionamento")
@login_required
def dashboard_comissionamento():
    user = current_user()
    filters = parse_commission_filters(request.args, user)
    data = fetch_commission_dashboard_data(user, filters)

    db = get_db()
    can_choose_seller = user["role"] in ("admin", "seller")
    sellers_sql = """
    SELECT id, username, full_name
    FROM users
    WHERE role = 'seller'
      AND is_active = 1
    """
    sellers_params = []
    if not is_master_user(user):
        owner_ids = get_user_scope_owner_ids(user)
        owner_clause, owner_params = build_in_clause("owner_admin_id", owner_ids)
        sellers_sql += f" AND {owner_clause}"
        sellers_params.extend(owner_params)
    sellers_sql += " ORDER BY username COLLATE NOCASE ASC"
    sellers = db.execute(sellers_sql, sellers_params).fetchall()

    companies = fetch_companies(user=user, include_inactive=False)
    courses = fetch_courses(user=user, include_inactive=False, seller_permission=None)
    community_admins = fetch_community_admins(user=user, include_inactive=False, include_masters=False)

    selected_seller = None
    if filters["seller_id"]:
        selected_seller = db.execute(
            """
            SELECT id, username, full_name
            FROM users
            WHERE id = ?
            """,
            (filters["seller_id"],),
        ).fetchone()

    return render_template(
        "comissionamento.html",
        filters=filters,
        can_choose_seller=can_choose_seller,
        sellers=sellers,
        companies=companies,
        courses=courses,
        community_admins=community_admins,
        selected_seller=selected_seller,
        summary=data["summary"],
        sales_rows=data["sales_rows"],
        rankings_courses=data["rankings_courses"],
        rankings_companies=data["rankings_companies"],
        rankings_sellers=data["rankings_sellers"],
        monthly_rows=data["monthly_rows"],
        best_course_by_sales=data["best_course_by_sales"],
        best_course_by_ticket=data["best_course_by_ticket"],
        favorite_company=data["favorite_company"],
        current_seller_rank=data["current_seller_rank"],
        commission_charts=data["charts"],
    )


def render_dashboard_view(recurring_only=False):
    user = current_user()
    filters = parse_filters(request.args, user)
    rows = fetch_installment_rows(filters, user, recurring_only=recurring_only)
    totals, charts = summarize_dashboard(rows)
    recurring_followup_rows = build_recurring_followup_queue(rows) if recurring_only else []

    companies = fetch_companies(user=user)
    courses = fetch_courses(user=user, seller_permission=None)
    sale_courses = fetch_courses(user=user, seller_permission="can_launch_sales")
    sellers = fetch_sellers(user=user)
    customer_names = fetch_customer_names(user)

    edit_sale = None
    edit_id_raw = (request.args.get("edit_sale") or "").strip()
    if edit_id_raw.isdigit():
        edit_id = int(edit_id_raw)
        db = get_db()
        sale = db.execute(
            """
            SELECT s.*
            FROM sales s
            WHERE s.id = ?
            """,
            (edit_id,),
        ).fetchone()
        if sale and can_edit_sale(user, sale["seller_id"], sale["owner_admin_id"], sale["course_id"]):
            edit_sale = dict(sale)
        elif sale:
            flash("Você não tem permissão para editar esta venda.", "error")

    sale_form = {
        "sale_id": edit_sale["id"] if edit_sale else "",
        "sale_date": edit_sale["sale_date"] if edit_sale else date.today().isoformat(),
        "customer_name": edit_sale["customer_name"] if edit_sale else "",
        "customer_phone": edit_sale["customer_phone"] if edit_sale and edit_sale["customer_phone"] else "",
        "customer_email": edit_sale["customer_email"] if edit_sale and edit_sale["customer_email"] else "",
        "company_id": edit_sale["company_id"] if edit_sale else "",
        "course_id": edit_sale["course_id"] if edit_sale else "",
        "payment_format": edit_sale["payment_format"] if edit_sale else "avista",
        "commission_payment_mode": edit_sale["commission_payment_mode"] if edit_sale else "per_installment",
        "installments_count": edit_sale["installments_count"] if edit_sale else 1,
        "total_value": edit_sale["total_value"] if edit_sale else "",
        "commission_percent": edit_sale["commission_percent"] if edit_sale else "",
        "seller_id": edit_sale["seller_id"] if edit_sale else (user["id"] if user["role"] == "seller" else ""),
        "notes": edit_sale["notes"] if edit_sale and edit_sale["notes"] else "",
    }

    filter_payload = filters_to_query(filters, user)
    if recurring_only:
        filter_payload["recurring_only"] = "1"
    export_url = url_for("export_xlsx", **filter_payload)
    current_path = request.full_path.rstrip("?")

    return render_template(
        "dashboard.html",
        filters=filters,
        rows=rows,
        totals=totals,
        charts=charts,
        companies=companies,
        courses=courses,
        sale_courses=sale_courses,
        sellers=sellers,
        customer_names=customer_names,
        sale_form=sale_form,
        can_write=user["role"] in ("admin", "seller"),
        export_url=export_url,
        current_path=current_path,
        recurring_only=recurring_only,
        recurring_followup_rows=recurring_followup_rows,
    )


@app.post("/sales/save")
@login_required
def save_sale():
    user = current_user()
    if user["role"] == "viewer":
        abort(403)
    validate_csrf()

    sale_id_raw = (request.form.get("sale_id") or "").strip()
    return_to = (request.form.get("return_to") or "").strip()

    try:
        sale_date = parse_date(request.form.get("sale_date"), "Data da venda")
        customer_name = clean_text(request.form.get("customer_name"), 120)
        if len(customer_name) < 2:
            raise ValueError("Informe um nome de cliente válido.")
        customer_phone = clean_text(request.form.get("customer_phone"), 40)
        customer_email = ""

        company_id = parse_int(request.form.get("company_id"), "Empresa")
        course_id = parse_int(request.form.get("course_id"), "Curso")
        payment_format = clean_text(request.form.get("payment_format"), 20)
        if payment_format not in PAYMENT_FORMATS:
            raise ValueError("Formato de pagamento inválido.")
        commission_payment_mode = clean_text(request.form.get("commission_payment_mode"), 40)
        if payment_format == "recorrencia":
            if commission_payment_mode not in COMMISSION_PAYMENT_MODES:
                raise ValueError("Confirme a modalidade da comissão para vendas recorrentes.")
        elif payment_format == "parcelado":
            commission_payment_mode = "upfront_first_installment"
        else:
            commission_payment_mode = "upfront_first_installment"
        total_value = parse_float(request.form.get("total_value"), "Valor total", minimum=0.01)
        commission_percent = parse_float(
            request.form.get("commission_percent"), "Percentual de comissão", minimum=0, maximum=100
        )
        installments_count = parse_int(request.form.get("installments_count"), "Quantidade de parcelas", minimum=1)
        if payment_format == "avista":
            installments_count = 1
        notes = clean_text(request.form.get("notes"), 1500)
    except ValueError as exc:
        flash(str(exc), "error")
        return redirect(url_for("dashboard"))

    db = get_db()
    company = db.execute(
        "SELECT id, owner_admin_id FROM companies WHERE id = ? AND is_active = 1",
        (company_id,),
    ).fetchone()
    if not company:
        flash("Empresa inválida ou inativa.", "error")
        return redirect(url_for("dashboard"))
    if not is_master_user(user) and not user_can_access_owner_scope(user, company["owner_admin_id"]):
        flash("Empresa fora da sua comunidade.", "error")
        return redirect(url_for("dashboard"))
    course = db.execute(
        "SELECT id FROM courses WHERE id = ? AND company_id = ? AND is_active = 1",
        (course_id, company_id),
    ).fetchone()
    if not course:
        flash("Curso inválido para a empresa selecionada.", "error")
        return redirect(url_for("dashboard"))

    if user["role"] == "admin":
        try:
            seller_id = parse_int(request.form.get("seller_id"), "Vendedor")
        except ValueError as exc:
            flash(str(exc), "error")
            return redirect(url_for("dashboard"))
        seller = db.execute(
            """
            SELECT id, role, is_active, owner_admin_id, is_master, is_manager
            FROM users
            WHERE id = ?
            """,
            (seller_id,),
        ).fetchone()
        if not seller or not seller["is_active"] or seller["role"] not in ("admin", "seller"):
            flash("Vendedor inválido.", "error")
            return redirect(url_for("dashboard"))
        if not is_master_user(user):
            seller_owner = seller["owner_admin_id"] or seller["id"]
            if not user_can_access_owner_scope(user, seller_owner):
                flash("Vendedor fora da sua comunidade.", "error")
                return redirect(url_for("dashboard"))
    else:
        seller_id = user["id"]
        seller = db.execute(
            "SELECT id, role, is_active, owner_admin_id, is_master, is_manager FROM users WHERE id = ?",
            (seller_id,),
        ).fetchone()
    if seller["role"] == "seller":
        if not seller_has_course_permission(seller["id"], course_id, "can_launch_sales"):
            flash("Este vendedor não tem permissão para lançar vendas neste curso.", "error")
            return redirect(url_for("dashboard"))
    sale_owner_admin_id = seller["owner_admin_id"] or (None if seller["is_master"] else seller["id"])

    installments = generate_installments(
        total_value=total_value,
        commission_percent=commission_percent,
        installments_count=installments_count,
        base_date=sale_date,
        payment_format=payment_format,
        commission_payment_mode=commission_payment_mode,
    )
    total_commission_expected = round(sum(item["commission_value"] for item in installments), 2)
    timestamp = now_iso()

    if sale_id_raw:
        if not sale_id_raw.isdigit():
            flash("Venda inválida para edição.", "error")
            return redirect(url_for("dashboard"))
        sale_id = int(sale_id_raw)
        existing = db.execute(
            "SELECT id, seller_id, owner_admin_id, course_id FROM sales WHERE id = ?",
            (sale_id,),
        ).fetchone()
        if not existing:
            flash("Venda não encontrada.", "error")
            return redirect(url_for("dashboard"))
        if not can_edit_sale(user, existing["seller_id"], existing["owner_admin_id"], existing["course_id"]):
            abort(403)

        db.execute(
            """
            UPDATE sales
            SET sale_date = ?,
                customer_name = ?,
                customer_phone = ?,
                customer_email = ?,
                company_id = ?,
                course_id = ?,
                seller_id = ?,
                owner_admin_id = ?,
                payment_format = ?,
                commission_payment_mode = ?,
                installments_count = ?,
                total_value = ?,
                commission_percent = ?,
                total_commission_expected = ?,
                notes = ?,
                updated_at = ?
            WHERE id = ?
            """,
            (
                sale_date,
                customer_name,
                customer_phone or None,
                customer_email or None,
                company_id,
                course_id,
                seller_id,
                sale_owner_admin_id,
                payment_format,
                commission_payment_mode,
                installments_count,
                total_value,
                commission_percent,
                total_commission_expected,
                notes or None,
                timestamp,
                sale_id,
            ),
        )
        db.execute("DELETE FROM sale_installments WHERE sale_id = ?", (sale_id,))
        new_sale_id = sale_id
        flash("Venda atualizada com sucesso.", "success")
    else:
        cursor = db.execute(
            """
            INSERT INTO sales (
              sale_date, customer_name, customer_phone, customer_email, company_id, course_id,
              seller_id, owner_admin_id, payment_format, commission_payment_mode, installments_count, total_value, commission_percent,
              total_commission_expected, notes, created_by, created_at, updated_at
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                sale_date,
                customer_name,
                customer_phone or None,
                customer_email or None,
                company_id,
                course_id,
                seller_id,
                sale_owner_admin_id,
                payment_format,
                commission_payment_mode,
                installments_count,
                total_value,
                commission_percent,
                total_commission_expected,
                notes or None,
                user["id"],
                timestamp,
                timestamp,
            ),
        )
        new_sale_id = cursor.lastrowid
        flash("Venda cadastrada com sucesso.", "success")

    default_status = "atrasado" if payment_format in ("recorrencia", "parcelado") else "confirmado"
    for item in installments:
        db.execute(
            """
            INSERT INTO sale_installments (
              sale_id, installment_number, due_date, month_key, installment_value,
              commission_value, status, paid_at, seller_followup_checked, seller_followup_checked_at, created_at, updated_at
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, NULL, 0, NULL, ?, ?)
            """,
            (
                new_sale_id,
                item["installment_number"],
                item["due_date"],
                item["month_key"],
                item["installment_value"],
                item["commission_value"],
                default_status,
                timestamp,
                timestamp,
            ),
        )

    db.commit()
    if valid_internal_return(return_to):
        return redirect(return_to)
    return redirect(url_for("dashboard"))


@app.post("/sales/<int:sale_id>/delete")
@login_required
def delete_sale(sale_id):
    user = current_user()
    if user["role"] == "viewer":
        abort(403)
    validate_csrf()

    db = get_db()
    sale = db.execute(
        "SELECT id, seller_id, owner_admin_id, course_id FROM sales WHERE id = ?",
        (sale_id,),
    ).fetchone()
    if not sale:
        flash("Venda não encontrada.", "error")
        return redirect(url_for("dashboard"))
    if not can_edit_sale(user, sale["seller_id"], sale["owner_admin_id"], sale["course_id"]):
        abort(403)

    db.execute("DELETE FROM sales WHERE id = ?", (sale_id,))
    db.commit()
    flash("Venda e parcelas removidas com sucesso.", "success")
    return redirect(url_for("dashboard"))


@app.post("/installments/<int:installment_id>/status")
@login_required
def update_installment_status(installment_id):
    user = current_user()
    if user["role"] == "viewer":
        abort(403)
    validate_csrf()
    return_to = (request.form.get("return_to") or "").strip()

    next_status = clean_text(request.form.get("status"), 20)
    if next_status not in INSTALLMENT_STATUSES:
        flash("Status de parcela inválido.", "error")
        return redirect(url_for("dashboard"))

    db = get_db()
    row = db.execute(
        """
        SELECT i.id, i.status, s.seller_id, s.owner_admin_id, s.course_id
        FROM sale_installments i
        INNER JOIN sales s ON s.id = i.sale_id
        WHERE i.id = ?
        """,
        (installment_id,),
    ).fetchone()
    if not row:
        flash("Parcela não encontrada.", "error")
        return redirect(url_for("dashboard"))
    if not can_edit_sale(user, row["seller_id"], row["owner_admin_id"], row["course_id"]):
        abort(403)

    paid_at = now_iso() if next_status == "confirmado" else None
    db.execute(
        """
        UPDATE sale_installments
        SET status = ?,
            paid_at = ?,
            seller_followup_checked = CASE WHEN ? = 'confirmado' THEN 1 ELSE seller_followup_checked END,
            seller_followup_checked_at = CASE WHEN ? = 'confirmado' THEN ? ELSE seller_followup_checked_at END,
            updated_at = ?
        WHERE id = ?
        """,
        (next_status, paid_at, next_status, next_status, now_iso(), now_iso(), installment_id),
    )
    db.commit()
    flash("Status da parcela atualizado.", "success")
    if valid_internal_return(return_to):
        return redirect(return_to)
    return redirect(url_for("dashboard"))


@app.post("/installments/<int:installment_id>/followup")
@login_required
def update_followup_status(installment_id):
    user = current_user()
    if user["role"] == "viewer":
        abort(403)
    validate_csrf()

    return_to = (request.form.get("return_to") or "").strip()
    action = clean_text(request.form.get("action"), 30)

    db = get_db()
    row = db.execute(
        """
        SELECT i.id, i.status, s.seller_id, s.owner_admin_id, s.course_id
        FROM sale_installments i
        INNER JOIN sales s ON s.id = i.sale_id
        WHERE i.id = ?
        """,
        (installment_id,),
    ).fetchone()
    if not row:
        flash("Parcela não encontrada.", "error")
        if valid_internal_return(return_to):
            return redirect(return_to)
        return redirect(url_for("dashboard_recorrencia"))
    if not can_edit_sale(user, row["seller_id"], row["owner_admin_id"], row["course_id"]):
        abort(403)

    now = now_iso()
    if action == "toggle_check":
        checked = request.form.get("checked") == "1"
        db.execute(
            """
            UPDATE sale_installments
            SET seller_followup_checked = ?,
                seller_followup_checked_at = ?,
                updated_at = ?
            WHERE id = ?
            """,
            (1 if checked else 0, now if checked else None, now, installment_id),
        )
        db.commit()
        flash("Checklist da recorrência atualizado.", "success")
    elif action == "mark_paid":
        db.execute(
            """
            UPDATE sale_installments
            SET status = 'confirmado',
                paid_at = ?,
                seller_followup_checked = 1,
                seller_followup_checked_at = ?,
                updated_at = ?
            WHERE id = ?
            """,
            (now, now, now, installment_id),
        )
        db.commit()
        flash("Parcela marcada como paga.", "success")
    else:
        flash("Ação inválida.", "error")

    if valid_internal_return(return_to):
        return redirect(return_to)
    return redirect(url_for("dashboard_recorrencia"))


@app.route("/admin/users", methods=["GET", "POST"])
@login_required
def manage_users():
    actor = current_user()
    if not is_admin_or_manager(actor):
        abort(403)

    db = get_db()
    actor_is_master = is_master_user(actor)
    actor_owner_ids = set(get_user_scope_owner_ids(actor))

    def resolve_target_owner(target):
        if target["is_master"]:
            return None
        return target["owner_admin_id"] or (target["id"] if target["role"] == "admin" else None)

    def can_manage_target(target):
        if actor_is_master:
            return True
        target_owner = resolve_target_owner(target)
        if target_owner is None or target_owner not in actor_owner_ids:
            return False
        if target["role"] == "admin" or target["is_master"]:
            return False
        if is_manager_user(actor):
            return not target["is_manager"]
        return True

    if request.method == "POST":
        validate_csrf()
        action = clean_text(request.form.get("action"), 30)

        if action == "create_simple_user":
            if not actor_is_master:
                flash("Somente o Administrador Mestre pode criar usuários diretos.", "error")
                return redirect(url_for("manage_users"))

            username = normalize_username(request.form.get("username"))
            full_name = clean_text(request.form.get("full_name"), 80)
            password = request.form.get("password") or ""
            confirm = request.form.get("password_confirm") or ""

            valid_user, user_message = validate_username(username)
            if not valid_user:
                flash(user_message, "error")
                return redirect(url_for("manage_users"))
            if len(full_name) < 3:
                flash("Nome completo muito curto.", "error")
                return redirect(url_for("manage_users"))
            valid_password, password_message = validate_password(password, confirm)
            if not valid_password:
                flash(password_message, "error")
                return redirect(url_for("manage_users"))

            exists = db.execute("SELECT id FROM users WHERE username = ?", (username,)).fetchone()
            if exists:
                flash("Usuário já existe.", "error")
                return redirect(url_for("manage_users"))

            db.execute(
                """
                INSERT INTO users (
                    username,
                    email,
                    full_name,
                    role,
                    password_hash,
                    password_plaintext,
                    header_label,
                    invited_by_username,
                    invited_by_email,
                    owner_admin_id,
                    is_master,
                    is_manager,
                    is_active,
                    created_at
                )
                VALUES (?, NULL, ?, 'viewer', ?, ?, NULL, ?, NULL, NULL, 0, 0, 1, ?)
                """,
                (
                    username,
                    full_name,
                    generate_password_hash(password),
                    password,
                    actor["username"],
                    now_iso(),
                ),
            )
            db.commit()
            flash("Usuário criado com sucesso. Você pode ajustar perfil e comunidade na tabela abaixo.", "success")
            return redirect(url_for("manage_users"))

        if action == "create_admin":
            if not actor_is_master:
                flash("Somente o Administrador Mestre pode criar administradores de comunidade.", "error")
                return redirect(url_for("manage_users"))

            username = normalize_username(request.form.get("username"))
            full_name = clean_text(request.form.get("full_name"), 80)
            password = request.form.get("password") or ""
            confirm = request.form.get("password_confirm") or ""

            valid_user, user_message = validate_username(username)
            if not valid_user:
                flash(user_message, "error")
                return redirect(url_for("manage_users"))
            if len(full_name) < 3:
                flash("Nome completo muito curto.", "error")
                return redirect(url_for("manage_users"))
            valid_password, password_message = validate_password(password, confirm)
            if not valid_password:
                flash(password_message, "error")
                return redirect(url_for("manage_users"))

            exists = db.execute("SELECT id FROM users WHERE username = ?", (username,)).fetchone()
            if exists:
                flash("Usuário já existe.", "error")
                return redirect(url_for("manage_users"))

            cursor = db.execute(
                """
                INSERT INTO users (
                    username,
                    email,
                    full_name,
                    role,
                    password_hash,
                    password_plaintext,
                    header_label,
                    invited_by_username,
                    invited_by_email,
                    owner_admin_id,
                    is_master,
                    is_manager,
                    is_active,
                    created_at
                )
                VALUES (?, NULL, ?, 'admin', ?, ?, NULL, ?, NULL, NULL, 0, 0, 1, ?)
                """,
                (
                    username,
                    full_name,
                    generate_password_hash(password),
                    password,
                    actor["username"],
                    now_iso(),
                ),
            )
            db.execute(
                "UPDATE users SET owner_admin_id = ? WHERE id = ?",
                (cursor.lastrowid, cursor.lastrowid),
            )
            db.execute(
                """
                INSERT INTO user_community_memberships (user_id, owner_admin_id, is_active, created_at, updated_at)
                VALUES (?, ?, 1, ?, ?)
                ON CONFLICT(user_id, owner_admin_id)
                DO UPDATE SET is_active = 1, updated_at = excluded.updated_at
                """,
                (cursor.lastrowid, cursor.lastrowid, now_iso(), now_iso()),
            )
            db.commit()
            flash("Administrador de comunidade criado com sucesso.", "success")
            return redirect(url_for("manage_users"))

        if action == "create_request":
            username = normalize_username(request.form.get("username"))
            full_name = clean_text(request.form.get("full_name"), 80)
            desired_profile_raw = clean_text(request.form.get("desired_profile"), 20).lower()
            referral_username = normalize_username(request.form.get("invite_username"))
            preferred_owner_admin_id = None
            if actor_is_master:
                try:
                    preferred_owner_admin_id = parse_int(
                        request.form.get("preferred_owner_admin_id"),
                        "Comunidade",
                        allow_empty=True,
                    )
                except ValueError as exc:
                    flash(str(exc), "error")
                    return redirect(url_for("manage_users"))
            else:
                preferred_owner_admin_id = None
            password = request.form.get("password") or ""
            confirm = request.form.get("password_confirm") or ""

            valid_user, user_message = validate_username(username)
            if not valid_user:
                flash(user_message, "error")
                return redirect(url_for("manage_users"))
            try:
                desired_profile, _ = normalize_requested_profile(desired_profile_raw)
            except ValueError:
                flash("Perfil solicitado inválido.", "error")
                return redirect(url_for("manage_users"))
            if desired_profile not in ("seller", "viewer"):
                flash("Perfil solicitado inválido.", "error")
                return redirect(url_for("manage_users"))
            if len(full_name) < 3:
                flash("Nome completo muito curto.", "error")
                return redirect(url_for("manage_users"))
            valid_password, password_message = validate_password(password, confirm)
            if not valid_password:
                flash(password_message, "error")
                return redirect(url_for("manage_users"))
            valid_ref, ref_message = validate_username(referral_username)
            if not valid_ref:
                flash("Usuário de indicação inválido.", "error")
                return redirect(url_for("manage_users"))

            exists = db.execute("SELECT id FROM users WHERE username = ?", (username,)).fetchone()
            if exists:
                flash("Usuário já existe.", "error")
                return redirect(url_for("manage_users"))
            pending_same = db.execute(
                "SELECT id FROM user_access_requests WHERE requested_username = ? AND status = 'pendente'",
                (username,),
            ).fetchone()
            if pending_same:
                flash("Já existe uma solicitação pendente para este usuário.", "error")
                return redirect(url_for("manage_users"))

            referral = db.execute(
                """
                SELECT id, username, role, is_master, is_manager, owner_admin_id, is_active
                FROM users
                WHERE username = ?
                """,
                (referral_username,),
            )
            referral = referral.fetchone()
            if not referral or not referral["is_active"]:
                flash("Usuário de indicação não encontrado ou inativo.", "error")
                return redirect(url_for("manage_users"))

            referral_is_valid = referral["role"] == "admin"
            if not referral_is_valid:
                flash("Usuário de indicação deve ser um Administrador ou Administrador Mestre.", "error")
                return redirect(url_for("manage_users"))

            owner_admin_id = resolve_owner_admin_for_referral(referral, preferred_owner_admin_id)
            request_note = None
            if owner_admin_id is None:
                request_note = "Comunidade pendente de definição no momento da aprovação."
            if not actor_is_master and owner_admin_id not in actor_owner_ids:
                flash("Você só pode criar solicitações para a sua comunidade.", "error")
                return redirect(url_for("manage_users"))

            db.execute(
                """
                INSERT INTO user_access_requests (
                    requested_username,
                    requested_full_name,
                    desired_profile,
                    password_hash,
                    password_plaintext,
                    referral_username,
                    owner_admin_id,
                    requested_community_id,
                    status,
                    request_note,
                    requested_at
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, NULL, 'pendente', ?, ?)
                """,
                (
                    username,
                    full_name,
                    desired_profile,
                    generate_password_hash(password),
                    password,
                    referral_username,
                    owner_admin_id,
                    request_note,
                    now_iso(),
                ),
            )
            db.commit()
            flash("Solicitação criada e enviada para aprovação.", "success")
            return redirect(url_for("manage_users"))

        if action == "save_user_changes":
            try:
                user_id = parse_int(request.form.get("user_id"), "Usuário")
            except ValueError as exc:
                flash(str(exc), "error")
                return redirect(url_for("manage_users"))

            target = db.execute(
                "SELECT id, username, full_name, role, is_manager, is_master, is_active, owner_admin_id FROM users WHERE id = ?",
                (user_id,),
            ).fetchone()
            if not target:
                flash("Usuário não encontrado.", "error")
                return redirect(url_for("manage_users"))
            if not can_manage_target(target):
                flash("Sem permissão para alterar este usuário.", "error")
                return redirect(url_for("manage_users"))

            desired_active_raw = (request.form.get("is_active") or "").strip()
            if desired_active_raw not in ("0", "1"):
                flash("Status inválido.", "error")
                return redirect(url_for("manage_users"))
            desired_active = 1 if desired_active_raw == "1" else 0

            if target["id"] == actor["id"] and target["is_active"] and desired_active == 0:
                flash("Não é possível desativar o próprio usuário conectado.", "error")
                return redirect(url_for("manage_users"))

            if actor_is_master:
                next_username = normalize_username(request.form.get("edit_username"))
                next_full_name = clean_text(request.form.get("edit_full_name"), 80)
                if not next_username:
                    next_username = target["username"]
                if not next_full_name:
                    next_full_name = target["full_name"]
                valid_user, user_message = validate_username(next_username)
                if not valid_user:
                    flash(user_message, "error")
                    return redirect(url_for("manage_users"))
                if len(next_full_name) < 3:
                    flash("Nome completo muito curto.", "error")
                    return redirect(url_for("manage_users"))
                if next_username != target["username"]:
                    duplicate = db.execute(
                        "SELECT id FROM users WHERE username = ? AND id <> ?",
                        (next_username, user_id),
                    ).fetchone()
                    if duplicate:
                        flash("Esse login já está em uso por outro usuário.", "error")
                        return redirect(url_for("manage_users"))
                profile_key = clean_text(request.form.get("profile"), 30).lower()
                if profile_key not in ("master", "admin", "seller", "viewer"):
                    flash("Perfil inválido.", "error")
                    return redirect(url_for("manage_users"))
                try:
                    community_admin_id = parse_int(
                        request.form.get("community_admin_id"),
                        "Comunidade",
                        allow_empty=True,
                    )
                except ValueError as exc:
                    flash(str(exc), "error")
                    return redirect(url_for("manage_users"))
            else:
                next_username = target["username"]
                next_full_name = target["full_name"]
                profile_key = profile_key_from_user_row(target)
                community_admin_id = target["owner_admin_id"]

            next_role = target["role"]
            next_is_manager = 0
            next_is_master = 1 if target["is_master"] else 0
            next_owner_admin_id = target["owner_admin_id"]

            if actor_is_master:
                next_role = "viewer"
                next_is_manager = 0
                next_is_master = 0
                next_owner_admin_id = None

                if profile_key == "master":
                    next_role = "admin"
                    next_is_master = 1
                elif profile_key == "admin":
                    next_role = "admin"
                    next_owner_admin_id = community_admin_id or user_id
                else:
                    next_role = "seller" if profile_key == "seller" else "viewer"
                    if community_admin_id is not None:
                        community_admin = db.execute(
                            """
                            SELECT id
                            FROM users
                            WHERE id = ?
                              AND role = 'admin'
                              AND COALESCE(is_master, 0) = 0
                              AND is_active = 1
                            """,
                            (community_admin_id,),
                        ).fetchone()
                        if not community_admin:
                            flash("Comunidade inválida.", "error")
                            return redirect(url_for("manage_users"))
                        next_owner_admin_id = community_admin["id"]
                    else:
                        next_owner_admin_id = None

                if next_role == "admin" and next_is_master == 0:
                    if next_owner_admin_id is None:
                        next_owner_admin_id = user_id
                if next_is_master == 1:
                    next_owner_admin_id = None
            else:
                if target["role"] == "admin" and not target["is_master"]:
                    next_owner_admin_id = target["id"]

            if target["role"] == "admin" and target["is_active"] and (next_role != "admin" or desired_active == 0):
                active_admins = db.execute(
                    "SELECT COUNT(*) AS total FROM users WHERE role = 'admin' AND is_active = 1"
                ).fetchone()["total"]
                if active_admins <= 1:
                    flash("Não é possível remover ou desativar o último admin ativo.", "error")
                    return redirect(url_for("manage_users"))

            if target["is_master"] and target["is_active"] and (next_is_master == 0 or desired_active == 0):
                active_masters = db.execute(
                    """
                    SELECT COUNT(*) AS total
                    FROM users
                    WHERE role = 'admin' AND COALESCE(is_master, 0) = 1 AND is_active = 1
                    """
                ).fetchone()["total"]
                if active_masters <= 1:
                    flash("Não é possível remover ou desativar o último mestre ativo.", "error")
                    return redirect(url_for("manage_users"))

            invited_by_reference = None
            if actor_is_master and next_role in ("seller", "viewer"):
                invited_by_reference = community_admin_id

            db.execute(
                """
                UPDATE users
                SET username = ?,
                    full_name = ?,
                    role = ?,
                    is_manager = ?,
                    is_master = ?,
                    owner_admin_id = ?,
                    is_active = ?,
                    invited_by_username = CASE
                        WHEN ? IS NULL THEN invited_by_username
                        ELSE (SELECT username FROM users WHERE id = ?)
                    END
                WHERE id = ?
                """,
                (
                    next_username,
                    next_full_name,
                    next_role,
                    next_is_manager,
                    next_is_master,
                    next_owner_admin_id,
                    desired_active,
                    invited_by_reference,
                    invited_by_reference,
                    user_id,
                ),
            )

            if next_owner_admin_id is not None:
                db.execute(
                    """
                    INSERT INTO user_community_memberships (user_id, owner_admin_id, is_active, created_at, updated_at)
                    VALUES (?, ?, 1, ?, ?)
                    ON CONFLICT(user_id, owner_admin_id)
                    DO UPDATE SET is_active = 1, updated_at = excluded.updated_at
                    """,
                    (user_id, next_owner_admin_id, now_iso(), now_iso()),
                )

            new_password = request.form.get("new_password") or ""
            if new_password:
                valid_password, password_message = validate_password(new_password)
                if not valid_password:
                    flash(password_message, "error")
                    db.rollback()
                    return redirect(url_for("manage_users"))
                db.execute(
                    "UPDATE users SET password_hash = ?, password_plaintext = ? WHERE id = ?",
                    (generate_password_hash(new_password), new_password, user_id),
                )

            db.commit()
            flash("Alterações do usuário salvas.", "success")
            return redirect(url_for("manage_users"))

        if action == "update_hierarchy":
            if not actor_is_master:
                flash("Somente o Administrador Mestre pode editar hierarquias.", "error")
                return redirect(url_for("manage_users"))
            try:
                user_id = parse_int(request.form.get("user_id"), "Usuário")
            except ValueError as exc:
                flash(str(exc), "error")
                return redirect(url_for("manage_users"))
            profile_key = clean_text(request.form.get("profile"), 30).lower()
            if profile_key not in ("master", "admin", "seller", "viewer"):
                flash("Perfil inválido.", "error")
                return redirect(url_for("manage_users"))
            try:
                community_admin_id = parse_int(
                    request.form.get("community_admin_id"),
                    "Comunidade",
                    allow_empty=True,
                )
            except ValueError as exc:
                flash(str(exc), "error")
                return redirect(url_for("manage_users"))
            target = db.execute(
                "SELECT id, username, role, is_manager, is_master, is_active FROM users WHERE id = ?",
                (user_id,),
            ).fetchone()
            if not target:
                flash("Usuário não encontrado.", "error")
                return redirect(url_for("manage_users"))

            if target["role"] == "admin" and profile_key not in ("admin", "master") and target["is_active"]:
                active_admins = db.execute(
                    "SELECT COUNT(*) AS total FROM users WHERE role = 'admin' AND is_active = 1"
                ).fetchone()["total"]
                if active_admins <= 1:
                    flash("Não é possível remover o último admin ativo.", "error")
                    return redirect(url_for("manage_users"))

            if target["is_master"] and profile_key != "master" and target["is_active"]:
                active_masters = db.execute(
                    """
                    SELECT COUNT(*) AS total
                    FROM users
                    WHERE role = 'admin' AND COALESCE(is_master, 0) = 1 AND is_active = 1
                    """
                ).fetchone()["total"]
                if active_masters <= 1:
                    flash("Não é possível remover o último mestre ativo.", "error")
                    return redirect(url_for("manage_users"))

            next_role = "viewer"
            next_is_manager = 0
            next_is_master = 0
            next_owner_admin_id = None

            if profile_key == "master":
                next_role = "admin"
                next_is_master = 1
                next_owner_admin_id = None
            elif profile_key == "admin":
                next_role = "admin"
                next_is_master = 0
                next_owner_admin_id = community_admin_id or user_id
            else:
                if profile_key == "seller":
                    next_role = "seller"
                else:
                    next_role = "viewer"
                if community_admin_id is not None:
                    community_admin = db.execute(
                        """
                        SELECT id, username
                        FROM users
                        WHERE id = ?
                          AND role = 'admin'
                          AND COALESCE(is_master, 0) = 0
                        """,
                        (community_admin_id,),
                    ).fetchone()
                    if not community_admin:
                        flash("Comunidade inválida.", "error")
                        return redirect(url_for("manage_users"))
                    next_owner_admin_id = community_admin["id"]
                else:
                    next_owner_admin_id = None

            if next_role == "admin":
                if next_is_master == 0 and next_owner_admin_id is None:
                    next_owner_admin_id = user_id

            db.execute(
                """
                UPDATE users
                SET role = ?,
                    is_manager = ?,
                    is_master = ?,
                    owner_admin_id = ?,
                    invited_by_username = CASE
                        WHEN ? IS NULL THEN invited_by_username
                        ELSE (SELECT username FROM users WHERE id = ?)
                    END
                WHERE id = ?
                """,
                (
                    next_role,
                    next_is_manager,
                    next_is_master,
                    next_owner_admin_id,
                    community_admin_id,
                    community_admin_id,
                    user_id,
                ),
            )
            if next_owner_admin_id is not None:
                db.execute(
                    """
                    INSERT INTO user_community_memberships (user_id, owner_admin_id, is_active, created_at, updated_at)
                    VALUES (?, ?, 1, ?, ?)
                    ON CONFLICT(user_id, owner_admin_id)
                    DO UPDATE SET is_active = 1, updated_at = excluded.updated_at
                    """,
                    (user_id, next_owner_admin_id, now_iso(), now_iso()),
                )
            db.commit()
            flash("Hierarquia atualizada.", "success")
            return redirect(url_for("manage_users"))

        if action == "toggle_active":
            try:
                user_id = parse_int(request.form.get("user_id"), "Usuário")
            except ValueError as exc:
                flash(str(exc), "error")
                return redirect(url_for("manage_users"))

            target = db.execute(
                "SELECT id, role, is_manager, is_master, is_active, owner_admin_id FROM users WHERE id = ?",
                (user_id,),
            ).fetchone()
            if not target:
                flash("Usuário não encontrado.", "error")
                return redirect(url_for("manage_users"))
            if target["id"] == current_user()["id"] and target["is_active"]:
                flash("Não é possível desativar o próprio usuário conectado.", "error")
                return redirect(url_for("manage_users"))
            if not can_manage_target(target):
                flash("Sem permissão para alterar este usuário.", "error")
                return redirect(url_for("manage_users"))
            if target["role"] == "admin" and target["is_active"]:
                active_admins = db.execute(
                    "SELECT COUNT(*) AS total FROM users WHERE role = 'admin' AND is_active = 1"
                ).fetchone()["total"]
                if active_admins <= 1:
                    flash("Não é possível desativar o último admin ativo.", "error")
                    return redirect(url_for("manage_users"))
            if target["is_master"] and target["is_active"]:
                active_masters = db.execute(
                    """
                    SELECT COUNT(*) AS total
                    FROM users
                    WHERE role = 'admin' AND COALESCE(is_master, 0) = 1 AND is_active = 1
                    """
                ).fetchone()["total"]
                if active_masters <= 1:
                    flash("Não é possível desativar o último mestre ativo.", "error")
                    return redirect(url_for("manage_users"))

            db.execute("UPDATE users SET is_active = ? WHERE id = ?", (0 if target["is_active"] else 1, user_id))
            db.commit()
            flash("Status do usuário atualizado.", "success")
            return redirect(url_for("manage_users"))

        if action == "reset_password":
            try:
                user_id = parse_int(request.form.get("user_id"), "Usuário")
            except ValueError as exc:
                flash(str(exc), "error")
                return redirect(url_for("manage_users"))
            target = db.execute(
                "SELECT id, role, is_manager, is_master, owner_admin_id FROM users WHERE id = ?",
                (user_id,),
            ).fetchone()
            if not target:
                flash("Usuário não encontrado.", "error")
                return redirect(url_for("manage_users"))
            if not can_manage_target(target):
                flash("Sem permissão para alterar este usuário.", "error")
                return redirect(url_for("manage_users"))

            new_password = request.form.get("new_password") or ""
            valid_password, password_message = validate_password(new_password)
            if not valid_password:
                flash(password_message, "error")
                return redirect(url_for("manage_users"))

            db.execute(
                "UPDATE users SET password_hash = ?, password_plaintext = ? WHERE id = ?",
                (generate_password_hash(new_password), new_password, user_id),
            )
            db.commit()
            flash("Senha atualizada.", "success")
            return redirect(url_for("manage_users"))

        if action == "set_seller_course_permission":
            if actor["role"] != "admin":
                flash("Somente administradores podem configurar permissões de cursos.", "error")
                return redirect(url_for("manage_users"))
            try:
                target_user_id = parse_int(request.form.get("target_user_id"), "Vendedor")
                course_id = parse_int(request.form.get("course_id"), "Curso")
            except ValueError as exc:
                flash(str(exc), "error")
                return redirect(url_for("manage_users"))
            target = db.execute(
                """
                SELECT id, role, is_manager, owner_admin_id
                FROM users
                WHERE id = ? AND is_active = 1
                """,
                (target_user_id,),
            ).fetchone()
            if not target or target["role"] != "seller" or target["is_manager"]:
                flash("Selecione um vendedor válido.", "error")
                return redirect(url_for("manage_users"))
            target_owner = target["owner_admin_id"]
            if not actor_is_master and target_owner not in actor_owner_ids:
                flash("Sem permissão para alterar vendedor de outra comunidade.", "error")
                return redirect(url_for("manage_users"))
            course = db.execute(
                """
                SELECT c.id, comp.owner_admin_id
                FROM courses c
                INNER JOIN companies comp ON comp.id = c.company_id
                WHERE c.id = ?
                """,
                (course_id,),
            ).fetchone()
            if not course:
                flash("Curso não encontrado.", "error")
                return redirect(url_for("manage_users"))
            if course["owner_admin_id"] != target_owner:
                flash("Curso e vendedor precisam estar na mesma comunidade.", "error")
                return redirect(url_for("manage_users"))
            if not actor_is_master and course["owner_admin_id"] not in actor_owner_ids:
                flash("Sem permissão para alterar cursos de outra comunidade.", "error")
                return redirect(url_for("manage_users"))

            can_launch_sales = 1 if request.form.get("can_launch_sales") == "1" else 0
            can_edit_sales = 1 if request.form.get("can_edit_sales") == "1" else 0
            can_edit_course = 1 if request.form.get("can_edit_course") == "1" else 0
            db.execute(
                """
                INSERT INTO seller_course_permissions (
                    user_id, course_id, can_launch_sales, can_edit_sales, can_edit_course, is_active, created_at, updated_at
                )
                VALUES (?, ?, ?, ?, ?, 1, ?, ?)
                ON CONFLICT(user_id, course_id)
                DO UPDATE SET
                    can_launch_sales = excluded.can_launch_sales,
                    can_edit_sales = excluded.can_edit_sales,
                    can_edit_course = excluded.can_edit_course,
                    is_active = 1,
                    updated_at = excluded.updated_at
                """,
                (
                    target_user_id,
                    course_id,
                    can_launch_sales,
                    can_edit_sales,
                    can_edit_course,
                    now_iso(),
                    now_iso(),
                ),
            )
            db.commit()
            flash("Permissão de curso atualizada para o vendedor.", "success")
            return redirect(url_for("manage_users"))

        if action == "remove_seller_course_permission":
            if actor["role"] != "admin":
                flash("Somente administradores podem remover permissões de cursos.", "error")
                return redirect(url_for("manage_users"))
            try:
                permission_id = parse_int(request.form.get("permission_id"), "Permissão")
            except ValueError as exc:
                flash(str(exc), "error")
                return redirect(url_for("manage_users"))
            permission_row = db.execute(
                """
                SELECT p.id, p.user_id, comp.owner_admin_id
                FROM seller_course_permissions p
                INNER JOIN courses c ON c.id = p.course_id
                INNER JOIN companies comp ON comp.id = c.company_id
                WHERE p.id = ?
                """,
                (permission_id,),
            ).fetchone()
            if not permission_row:
                flash("Permissão não encontrada.", "error")
                return redirect(url_for("manage_users"))
            if not actor_is_master and permission_row["owner_admin_id"] not in actor_owner_ids:
                flash("Sem permissão para remover curso de outra comunidade.", "error")
                return redirect(url_for("manage_users"))
            db.execute("DELETE FROM seller_course_permissions WHERE id = ?", (permission_id,))
            db.commit()
            flash("Permissão de curso removida.", "success")
            return redirect(url_for("manage_users"))

        if action == "add_user_community_membership":
            if not actor_is_master:
                flash("Somente o Administrador Mestre pode gerenciar múltiplas comunidades.", "error")
                return redirect(url_for("manage_users"))
            try:
                user_id = parse_int(request.form.get("user_id"), "Usuário")
                owner_admin_id = parse_int(request.form.get("owner_admin_id"), "Comunidade")
            except ValueError as exc:
                flash(str(exc), "error")
                return redirect(url_for("manage_users"))
            target_user = db.execute(
                "SELECT id, role, is_master, owner_admin_id FROM users WHERE id = ?",
                (user_id,),
            ).fetchone()
            if not target_user:
                flash("Usuário não encontrado.", "error")
                return redirect(url_for("manage_users"))
            owner_admin = db.execute(
                """
                SELECT id
                FROM users
                WHERE id = ?
                  AND role = 'admin'
                  AND is_active = 1
                """,
                (owner_admin_id,),
            ).fetchone()
            if not owner_admin:
                flash("Comunidade inválida.", "error")
                return redirect(url_for("manage_users"))
            db.execute(
                """
                INSERT INTO user_community_memberships (user_id, owner_admin_id, is_active, created_at, updated_at)
                VALUES (?, ?, 1, ?, ?)
                ON CONFLICT(user_id, owner_admin_id)
                DO UPDATE SET is_active = 1, updated_at = excluded.updated_at
                """,
                (user_id, owner_admin_id, now_iso(), now_iso()),
            )
            if target_user["role"] != "admin" and target_user["owner_admin_id"] is None:
                db.execute("UPDATE users SET owner_admin_id = ? WHERE id = ?", (owner_admin_id, user_id))
            db.commit()
            flash("Vinculação de comunidade adicionada.", "success")
            return redirect(url_for("manage_users"))

        if action == "remove_user_community_membership":
            if not actor_is_master:
                flash("Somente o Administrador Mestre pode gerenciar múltiplas comunidades.", "error")
                return redirect(url_for("manage_users"))
            try:
                membership_id = parse_int(request.form.get("membership_id"), "Vinculação")
            except ValueError as exc:
                flash(str(exc), "error")
                return redirect(url_for("manage_users"))
            membership = db.execute(
                """
                SELECT id, user_id, owner_admin_id
                FROM user_community_memberships
                WHERE id = ?
                """,
                (membership_id,),
            ).fetchone()
            if not membership:
                flash("Vinculação não encontrada.", "error")
                return redirect(url_for("manage_users"))
            db.execute(
                """
                UPDATE user_community_memberships
                SET is_active = 0, updated_at = ?
                WHERE id = ?
                """,
                (now_iso(), membership_id),
            )
            db.commit()
            flash("Vinculação de comunidade removida.", "success")
            return redirect(url_for("manage_users"))

    users_sql = """
    SELECT
      u.id,
      u.username,
      u.full_name,
      u.role,
      u.is_manager,
      u.is_master,
      u.password_plaintext,
      u.invited_by_username,
      u.owner_admin_id,
      u.is_active,
      u.created_at,
      owner.full_name AS owner_admin_name,
      owner.username AS owner_admin_username
    FROM users u
    LEFT JOIN users owner ON owner.id = u.owner_admin_id
    """
    params = []
    if not actor_is_master:
        owner_clause, owner_params = build_in_clause(
            "COALESCE(NULLIF(u.owner_admin_id, 0), CASE WHEN u.role = 'admin' AND COALESCE(u.is_master, 0) = 0 THEN u.id END)",
            actor_owner_ids,
        )
        users_sql += f"""
        WHERE (
            {owner_clause}
            OR u.id = ?
        )
        """
        params.extend(owner_params)
        params.append(actor["id"])
    users_sql += " ORDER BY u.created_at DESC"
    users = db.execute(users_sql, params).fetchall()
    community_admins = fetch_community_admins(user=actor, include_inactive=True)
    membership_target_users = []
    user_memberships = []
    if actor_is_master:
        membership_target_users = db.execute(
            """
            SELECT id, full_name, username, role, is_master
            FROM users
            WHERE is_active = 1
            ORDER BY full_name COLLATE NOCASE ASC
            """
        ).fetchall()
        user_memberships = db.execute(
            """
            SELECT
              m.id,
              m.user_id,
              m.owner_admin_id,
              u.full_name AS user_name,
              u.username AS user_username,
              owner.full_name AS owner_name,
              owner.username AS owner_username
            FROM user_community_memberships m
            INNER JOIN users u ON u.id = m.user_id
            INNER JOIN users owner ON owner.id = m.owner_admin_id
            WHERE m.is_active = 1
            ORDER BY u.full_name COLLATE NOCASE ASC, owner.full_name COLLATE NOCASE ASC
            """
        ).fetchall()
    can_manage_seller_permissions = actor["role"] == "admin"
    manageable_sellers = []
    manageable_courses = []
    seller_course_permissions = []
    if can_manage_seller_permissions:
        sellers_sql = """
        SELECT id, full_name, username, owner_admin_id
        FROM users
        WHERE role = 'seller'
          AND COALESCE(is_manager, 0) = 0
          AND is_active = 1
        """
        sellers_params = []
        courses_sql = """
        SELECT c.id, c.name, comp.name AS company_name, comp.owner_admin_id
        FROM courses c
        INNER JOIN companies comp ON comp.id = c.company_id
        WHERE c.is_active = 1
        """
        courses_params = []
        permissions_sql = """
        SELECT
          p.id,
          p.user_id,
          p.course_id,
          p.can_launch_sales,
          p.can_edit_sales,
          p.can_edit_course,
          u.username AS seller_name,
          c.name AS course_name,
          comp.name AS company_name,
          comp.owner_admin_id
        FROM seller_course_permissions p
        INNER JOIN users u ON u.id = p.user_id
        INNER JOIN courses c ON c.id = p.course_id
        INNER JOIN companies comp ON comp.id = c.company_id
        WHERE p.is_active = 1
        """
        permissions_params = []
        if not actor_is_master:
            owner_clause, owner_params = build_in_clause("owner_admin_id", actor_owner_ids)
            sellers_sql += f" AND {owner_clause}"
            sellers_params.extend(owner_params)
            owner_clause_courses, owner_params_courses = build_in_clause("comp.owner_admin_id", actor_owner_ids)
            courses_sql += f" AND {owner_clause_courses}"
            courses_params.extend(owner_params_courses)
            owner_clause_permissions, owner_params_permissions = build_in_clause("comp.owner_admin_id", actor_owner_ids)
            permissions_sql += f" AND {owner_clause_permissions}"
            permissions_params.extend(owner_params_permissions)
        sellers_sql += " ORDER BY username COLLATE NOCASE ASC"
        courses_sql += " ORDER BY comp.name COLLATE NOCASE ASC, c.name COLLATE NOCASE ASC"
        permissions_sql += " ORDER BY seller_name COLLATE NOCASE ASC, company_name COLLATE NOCASE ASC, course_name COLLATE NOCASE ASC"
        manageable_sellers = db.execute(sellers_sql, sellers_params).fetchall()
        manageable_courses = db.execute(courses_sql, courses_params).fetchall()
        seller_course_permissions = db.execute(permissions_sql, permissions_params).fetchall()
    return render_template(
        "users.html",
        users=users,
        actor_is_master=actor_is_master,
        can_export_credentials=actor_is_master,
        community_admins=community_admins,
        membership_target_users=membership_target_users,
        user_memberships=user_memberships,
        can_manage_seller_permissions=can_manage_seller_permissions,
        manageable_sellers=manageable_sellers,
        manageable_courses=manageable_courses,
        seller_course_permissions=seller_course_permissions,
    )


@app.get("/admin/users/export-credenciais")
@login_required
def export_user_credentials():
    actor = current_user()
    if not is_master_user(actor):
        abort(403)
    if Workbook is None:
        flash("Instale openpyxl para exportar credenciais.", "error")
        return redirect(url_for("manage_users"))

    db = get_db()
    rows = db.execute(
        """
        SELECT full_name, username, COALESCE(password_plaintext, '') AS password_plaintext
        FROM users
        ORDER BY full_name COLLATE NOCASE ASC
        """
    ).fetchall()

    wb = Workbook()
    ws = wb.active
    ws.title = "Credenciais"
    ws.append(["Nome completo", "Login", "Senha"])
    for row in rows:
        ws.append(
            [
                row["full_name"],
                row["username"],
                row["password_plaintext"] or "Não disponível (defina nova senha)",
            ]
        )

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f"credenciais_usuarios_{date.today().isoformat()}.xlsx"
    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/admin/user-requests", methods=["GET", "POST"])
@login_required
def user_access_requests():
    actor = current_user()
    if not is_admin_or_manager(actor):
        abort(403)

    db = get_db()
    actor_is_master = is_master_user(actor)
    actor_owner_ids = set(get_user_scope_owner_ids(actor))

    if request.method == "POST":
        validate_csrf()
        action = clean_text(request.form.get("action"), 20)
        try:
            request_id = parse_int(request.form.get("request_id"), "Solicitação")
        except ValueError as exc:
            flash(str(exc), "error")
            return redirect(url_for("user_access_requests"))

        req = db.execute(
            """
            SELECT *
            FROM user_access_requests
            WHERE id = ?
            """,
            (request_id,),
        ).fetchone()
        if not req:
            flash("Solicitação não encontrada.", "error")
            return redirect(url_for("user_access_requests"))
        if req["status"] != "pendente":
            flash("Esta solicitação já foi analisada.", "error")
            return redirect(url_for("user_access_requests"))
        if not actor_is_master:
            owner_matches_scope = req["owner_admin_id"] in actor_owner_ids if req["owner_admin_id"] is not None else False
            referral_matches_actor = req["referral_username"] == actor["username"]
            if not owner_matches_scope and not referral_matches_actor:
                flash("Sem permissão para analisar solicitação de outra comunidade.", "error")
                return redirect(url_for("user_access_requests"))

        if action == "deny":
            db.execute(
                """
                UPDATE user_access_requests
                SET status = 'recusado',
                    reviewed_at = ?,
                    reviewed_by = ?,
                    decision_note = ?
                WHERE id = ?
                """,
                (now_iso(), actor["id"], clean_text(request.form.get("decision_note"), 200), request_id),
            )
            db.commit()
            flash("Solicitação recusada.", "success")
            return redirect(url_for("user_access_requests"))

        if action == "approve":
            role, is_manager = normalize_requested_profile(req["desired_profile"])
            exists = db.execute("SELECT id FROM users WHERE username = ?", (req["requested_username"],)).fetchone()
            if exists:
                flash("Usuário já existe. Solicitação recusada automaticamente.", "error")
                db.execute(
                    """
                    UPDATE user_access_requests
                    SET status = 'recusado', reviewed_at = ?, reviewed_by = ?, decision_note = 'Usuário já existe'
                    WHERE id = ?
                    """,
                    (now_iso(), actor["id"], request_id),
                )
                db.commit()
                return redirect(url_for("user_access_requests"))

            owner_admin_id = req["owner_admin_id"]
            try:
                selected_owner_admin_id = parse_int(
                    request.form.get("owner_admin_id"),
                    "Comunidade",
                    allow_empty=True,
                )
            except ValueError as exc:
                flash(str(exc), "error")
                return redirect(url_for("user_access_requests"))
            if selected_owner_admin_id is not None:
                owner_admin = db.execute(
                    """
                    SELECT id
                    FROM users
                    WHERE id = ?
                      AND role = 'admin'
                      AND is_active = 1
                    """,
                    (selected_owner_admin_id,),
                ).fetchone()
                if not owner_admin:
                    flash("Comunidade de destino inválida.", "error")
                    return redirect(url_for("user_access_requests"))
                owner_admin_id = owner_admin["id"]
            if not actor_is_master:
                if owner_admin_id is None:
                    owner_admin_id = get_owner_admin_id_for_user(actor)
                if owner_admin_id not in actor_owner_ids:
                    flash("Sem permissão para aprovar para esta comunidade.", "error")
                    return redirect(url_for("user_access_requests"))

            user_cursor = db.execute(
                """
                INSERT INTO users (
                    username, full_name, role, password_hash, password_plaintext, header_label,
                    invited_by_username, invited_by_email, owner_admin_id,
                    is_master, is_manager, is_active, created_at
                )
                VALUES (?, ?, ?, ?, ?, NULL, ?, NULL, ?, 0, ?, 1, ?)
                """,
                (
                    req["requested_username"],
                    req["requested_full_name"],
                    role,
                    req["password_hash"],
                    req["password_plaintext"],
                    req["referral_username"],
                    owner_admin_id,
                    is_manager,
                    now_iso(),
                ),
            )
            if owner_admin_id is not None:
                db.execute(
                    """
                    INSERT INTO user_community_memberships (user_id, owner_admin_id, is_active, created_at, updated_at)
                    VALUES (?, ?, 1, ?, ?)
                    ON CONFLICT(user_id, owner_admin_id)
                    DO UPDATE SET is_active = 1, updated_at = excluded.updated_at
                    """,
                    (user_cursor.lastrowid, owner_admin_id, now_iso(), now_iso()),
                )
            if role == "seller" and not is_manager and owner_admin_id is not None:
                seed_seller_course_permissions(
                    user_cursor.lastrowid,
                    owner_admin_id,
                    can_launch_sales=0,
                    can_edit_sales=0,
                    can_edit_course=0,
                )
            db.execute(
                """
                UPDATE user_access_requests
                SET status = 'aprovado',
                    reviewed_at = ?,
                    reviewed_by = ?,
                    owner_admin_id = ?,
                    decision_note = ?
                WHERE id = ?
                """,
                (
                    now_iso(),
                    actor["id"],
                    owner_admin_id,
                    clean_text(request.form.get("decision_note"), 200),
                    request_id,
                ),
            )
            db.commit()
            flash("Solicitação aprovada e usuário criado.", "success")
            return redirect(url_for("user_access_requests"))

    sql = """
    SELECT
      r.*,
      reviewer.full_name AS reviewer_name,
      owner.username AS owner_admin_username,
      owner.full_name AS owner_admin_full_name
    FROM user_access_requests r
    LEFT JOIN users reviewer ON reviewer.id = r.reviewed_by
    LEFT JOIN users owner ON owner.id = r.owner_admin_id
    """
    params = []
    if not actor_is_master:
        owner_clause, owner_params = build_in_clause("r.owner_admin_id", actor_owner_ids)
        sql += f" WHERE ({owner_clause} OR r.referral_username = ?)"
        params.extend(owner_params)
        params.append(actor["username"])
    sql += " ORDER BY CASE r.status WHEN 'pendente' THEN 0 WHEN 'aprovado' THEN 1 ELSE 2 END, r.requested_at DESC"
    requests_rows = db.execute(sql, params).fetchall()
    community_admins = fetch_community_admins(user=actor, include_inactive=False, include_masters=False)
    return render_template(
        "user_access_requests.html",
        requests=requests_rows,
        community_admins=community_admins,
        actor_is_master=actor_is_master,
    )


@app.get("/admin/master-features")
@login_required
def master_features():
    user = current_user()
    if not is_master_user(user):
        abort(403)

    db = get_db()
    totals = db.execute(
        """
        SELECT
          (SELECT COUNT(*) FROM users WHERE is_active = 1) AS active_users,
          (SELECT COUNT(*) FROM users WHERE role = 'admin' AND COALESCE(is_master, 0) = 1 AND is_active = 1) AS active_masters,
          (SELECT COUNT(*) FROM users WHERE role = 'admin' AND COALESCE(is_master, 0) = 0 AND is_active = 1) AS active_community_admins,
          (SELECT COUNT(*) FROM users WHERE role = 'seller' AND is_active = 1) AS active_sellers,
          (SELECT COUNT(*) FROM communities WHERE is_active = 1) AS active_communities,
          (SELECT COUNT(*) FROM sales) AS total_sales,
          (SELECT COUNT(*) FROM courses WHERE is_active = 1) AS active_courses,
          (SELECT COUNT(*) FROM user_access_requests WHERE status = 'pendente') AS pending_user_requests
        """
    ).fetchone()
    communities = fetch_communities(user=user, include_inactive=True)
    return render_template("master_features.html", totals=totals, communities=communities)


@app.route("/admin/communities", methods=["GET", "POST"])
@login_required
def manage_communities():
    actor = current_user()
    if not is_master_user(actor):
        abort(403)
    db = get_db()

    if request.method == "POST":
        validate_csrf()
        action = clean_text(request.form.get("action"), 30)

        if action == "create":
            name = clean_text(request.form.get("name"), 120)
            if len(name) < 2:
                flash("Nome da comunidade inválido.", "error")
                return redirect(url_for("manage_communities"))
            try:
                owner_admin_id = parse_int(
                    request.form.get("owner_admin_id"),
                    "Administrador da comunidade",
                    allow_empty=True,
                )
            except ValueError as exc:
                flash(str(exc), "error")
                return redirect(url_for("manage_communities"))

            if owner_admin_id is not None:
                owner_admin = db.execute(
                    """
                    SELECT id
                    FROM users
                    WHERE id = ?
                      AND role = 'admin'
                      AND COALESCE(is_master, 0) = 0
                      AND is_active = 1
                    """,
                    (owner_admin_id,),
                ).fetchone()
                if not owner_admin:
                    flash("Administrador da comunidade inválido.", "error")
                    return redirect(url_for("manage_communities"))

            now = now_iso()
            db.execute(
                """
                INSERT INTO communities (name, owner_admin_id, manager_user_id, is_active, created_at, updated_at)
                VALUES (?, ?, ?, 1, ?, ?)
                """,
                (name, owner_admin_id, None, now, now),
            )
            if owner_admin_id is not None:
                db.execute(
                    """
                    INSERT INTO user_community_memberships (user_id, owner_admin_id, is_active, created_at, updated_at)
                    VALUES (?, ?, 1, ?, ?)
                    ON CONFLICT(user_id, owner_admin_id)
                    DO UPDATE SET is_active = 1, updated_at = excluded.updated_at
                    """,
                    (owner_admin_id, owner_admin_id, now, now),
                )
            db.commit()
            flash("Comunidade criada.", "success")
            return redirect(url_for("manage_communities"))

        if action == "save_changes":
            try:
                community_id = parse_int(request.form.get("community_id"), "Comunidade")
            except ValueError as exc:
                flash(str(exc), "error")
                return redirect(url_for("manage_communities"))
            name = clean_text(request.form.get("name"), 120)
            try:
                owner_admin_id = parse_int(
                    request.form.get("owner_admin_id"),
                    "Administrador da comunidade",
                    allow_empty=True,
                )
            except ValueError as exc:
                flash(str(exc), "error")
                return redirect(url_for("manage_communities"))
            try:
                manager_user_id = parse_int(
                    request.form.get("manager_user_id"),
                    "Responsável comercial",
                    allow_empty=True,
                )
            except ValueError as exc:
                flash(str(exc), "error")
                return redirect(url_for("manage_communities"))
            desired_active_raw = (request.form.get("is_active") or "").strip()
            if desired_active_raw not in ("0", "1"):
                flash("Status da comunidade inválido.", "error")
                return redirect(url_for("manage_communities"))
            desired_active = 1 if desired_active_raw == "1" else 0

            community = db.execute("SELECT id, owner_admin_id FROM communities WHERE id = ?", (community_id,)).fetchone()
            if not community:
                flash("Comunidade não encontrada.", "error")
                return redirect(url_for("manage_communities"))
            if len(name) < 2:
                flash("Nome da comunidade inválido.", "error")
                return redirect(url_for("manage_communities"))
            if owner_admin_id is not None:
                owner_admin = db.execute(
                    """
                    SELECT id
                    FROM users
                    WHERE id = ?
                      AND role = 'admin'
                      AND COALESCE(is_master, 0) = 0
                      AND is_active = 1
                    """,
                    (owner_admin_id,),
                ).fetchone()
                if not owner_admin:
                    flash("Administrador da comunidade inválido.", "error")
                    return redirect(url_for("manage_communities"))
            if manager_user_id is not None:
                manager = db.execute(
                    """
                    SELECT id
                    FROM users
                    WHERE id = ?
                      AND role = 'seller'
                      AND COALESCE(is_manager, 0) = 1
                      AND owner_admin_id = ?
                    """,
                    (manager_user_id, owner_admin_id),
                ).fetchone()
                if not manager:
                    flash("Responsável comercial inválido para esta comunidade.", "error")
                    return redirect(url_for("manage_communities"))

            db.execute(
                """
                UPDATE communities
                SET name = ?, owner_admin_id = ?, manager_user_id = ?, is_active = ?, updated_at = ?
                WHERE id = ?
                """,
                (name, owner_admin_id, manager_user_id, desired_active, now_iso(), community_id),
            )
            if owner_admin_id is not None:
                db.execute(
                    """
                    INSERT INTO user_community_memberships (user_id, owner_admin_id, is_active, created_at, updated_at)
                    VALUES (?, ?, 1, ?, ?)
                    ON CONFLICT(user_id, owner_admin_id)
                    DO UPDATE SET is_active = 1, updated_at = excluded.updated_at
                    """,
                    (owner_admin_id, owner_admin_id, now_iso(), now_iso()),
                )
            db.commit()
            flash("Alterações da comunidade salvas.", "success")
            return redirect(url_for("manage_communities"))

        if action == "update":
            try:
                community_id = parse_int(request.form.get("community_id"), "Comunidade")
            except ValueError as exc:
                flash(str(exc), "error")
                return redirect(url_for("manage_communities"))
            name = clean_text(request.form.get("name"), 120)
            try:
                owner_admin_id = parse_int(
                    request.form.get("owner_admin_id"),
                    "Administrador da comunidade",
                    allow_empty=True,
                )
            except ValueError as exc:
                flash(str(exc), "error")
                return redirect(url_for("manage_communities"))
            try:
                manager_user_id = parse_int(
                    request.form.get("manager_user_id"),
                    "Responsável comercial",
                    allow_empty=True,
                )
            except ValueError as exc:
                flash(str(exc), "error")
                return redirect(url_for("manage_communities"))
            community = db.execute("SELECT id, owner_admin_id FROM communities WHERE id = ?", (community_id,)).fetchone()
            if not community:
                flash("Comunidade não encontrada.", "error")
                return redirect(url_for("manage_communities"))
            if len(name) < 2:
                flash("Nome da comunidade inválido.", "error")
                return redirect(url_for("manage_communities"))
            if owner_admin_id is not None:
                owner_admin = db.execute(
                    """
                    SELECT id
                    FROM users
                    WHERE id = ?
                      AND role = 'admin'
                      AND COALESCE(is_master, 0) = 0
                      AND is_active = 1
                    """,
                    (owner_admin_id,),
                ).fetchone()
                if not owner_admin:
                    flash("Administrador da comunidade inválido.", "error")
                    return redirect(url_for("manage_communities"))
            if manager_user_id is not None:
                manager = db.execute(
                    """
                    SELECT id
                    FROM users
                    WHERE id = ?
                      AND role = 'seller'
                      AND COALESCE(is_manager, 0) = 1
                      AND owner_admin_id = ?
                    """,
                    (manager_user_id, owner_admin_id),
                ).fetchone()
                if not manager:
                    flash("Responsável comercial inválido para esta comunidade.", "error")
                    return redirect(url_for("manage_communities"))

            db.execute(
                """
                UPDATE communities
                SET name = ?, owner_admin_id = ?, manager_user_id = ?, updated_at = ?
                WHERE id = ?
                """,
                (name, owner_admin_id, manager_user_id, now_iso(), community_id),
            )
            db.commit()
            flash("Comunidade atualizada.", "success")
            return redirect(url_for("manage_communities"))

        if action == "toggle_active":
            try:
                community_id = parse_int(request.form.get("community_id"), "Comunidade")
            except ValueError as exc:
                flash(str(exc), "error")
                return redirect(url_for("manage_communities"))
            row = db.execute("SELECT id, is_active FROM communities WHERE id = ?", (community_id,)).fetchone()
            if not row:
                flash("Comunidade não encontrada.", "error")
                return redirect(url_for("manage_communities"))
            db.execute(
                "UPDATE communities SET is_active = ?, updated_at = ? WHERE id = ?",
                (0 if row["is_active"] else 1, now_iso(), community_id),
            )
            db.commit()
            flash("Status da comunidade atualizado.", "success")
            return redirect(url_for("manage_communities"))

    communities = fetch_communities(user=actor, include_inactive=True)
    community_admins = fetch_community_admins(user=actor, include_inactive=False, include_masters=False)
    managers = db.execute(
        """
        SELECT id, full_name, username, owner_admin_id
        FROM users
        WHERE role = 'seller' AND COALESCE(is_manager, 0) = 1 AND is_active = 1
        ORDER BY username COLLATE NOCASE ASC
        """
    ).fetchall()
    return render_template(
        "communities.html",
        communities=communities,
        community_admins=community_admins,
        managers=managers,
    )


@app.route("/admin/companies", methods=["GET", "POST"])
@login_required
def manage_companies():
    db = get_db()
    actor = current_user()
    if not is_admin_or_manager(actor):
        abort(403)
    actor_is_master = is_master_user(actor)
    actor_owner_admin_id = get_owner_admin_id_for_user(actor)
    actor_owner_ids = set(get_user_scope_owner_ids(actor))

    if request.method == "POST":
        validate_csrf()
        action = clean_text(request.form.get("action"), 30)

        if action == "create":
            name = clean_text(request.form.get("name"), 120)
            if len(name) < 2:
                flash("Nome da empresa inválido.", "error")
                return redirect(url_for("manage_companies"))
            owner_admin_id = actor_owner_admin_id
            if not actor_is_master:
                try:
                    selected_owner = parse_int(
                        request.form.get("owner_admin_id"),
                        "Comunidade",
                        allow_empty=True,
                    )
                except ValueError as exc:
                    flash(str(exc), "error")
                    return redirect(url_for("manage_companies"))
                if selected_owner is not None:
                    owner_admin_id = selected_owner
                if owner_admin_id not in actor_owner_ids:
                    flash("Comunidade inválida para criação da empresa.", "error")
                    return redirect(url_for("manage_companies"))
            if actor_is_master:
                try:
                    owner_admin_id = parse_int(
                        request.form.get("owner_admin_id"),
                        "Comunidade",
                        allow_empty=True,
                    )
                except ValueError as exc:
                    flash(str(exc), "error")
                    return redirect(url_for("manage_companies"))
                if owner_admin_id is None:
                    flash("Selecione a comunidade dona da empresa.", "error")
                    return redirect(url_for("manage_companies"))
                owner_row = db.execute(
                    """
                    SELECT id
                    FROM users
                    WHERE id = ?
                      AND role = 'admin'
                      AND COALESCE(is_master, 0) = 0
                    """,
                    (owner_admin_id,),
                ).fetchone()
                if not owner_row:
                    flash("Comunidade inválida.", "error")
                    return redirect(url_for("manage_companies"))
            exists = db.execute("SELECT id FROM companies WHERE lower(name) = lower(?)", (name,)).fetchone()
            if exists:
                flash("Empresa já existe.", "error")
                return redirect(url_for("manage_companies"))
            now = now_iso()
            db.execute(
                "INSERT INTO companies (name, owner_admin_id, is_active, created_at, updated_at) VALUES (?, ?, 1, ?, ?)",
                (name, owner_admin_id, now, now),
            )
            db.commit()
            flash("Empresa criada.", "success")
            return redirect(url_for("manage_companies"))

        if action == "update":
            try:
                company_id = parse_int(request.form.get("company_id"), "Empresa")
            except ValueError as exc:
                flash(str(exc), "error")
                return redirect(url_for("manage_companies"))
            target = db.execute("SELECT id, owner_admin_id FROM companies WHERE id = ?", (company_id,)).fetchone()
            if not target:
                flash("Empresa não encontrada.", "error")
                return redirect(url_for("manage_companies"))
            if not actor_is_master and target["owner_admin_id"] not in actor_owner_ids:
                flash("Sem permissão para editar empresa de outra comunidade.", "error")
                return redirect(url_for("manage_companies"))
            name = clean_text(request.form.get("name"), 120)
            if len(name) < 2:
                flash("Nome da empresa inválido.", "error")
                return redirect(url_for("manage_companies"))
            duplicate = db.execute(
                "SELECT id FROM companies WHERE lower(name) = lower(?) AND id <> ?",
                (name, company_id),
            ).fetchone()
            if duplicate:
                flash("Já existe outra empresa com este nome.", "error")
                return redirect(url_for("manage_companies"))
            db.execute("UPDATE companies SET name = ?, updated_at = ? WHERE id = ?", (name, now_iso(), company_id))
            db.commit()
            flash("Empresa atualizada.", "success")
            return redirect(url_for("manage_companies"))

        if action == "toggle_active":
            try:
                company_id = parse_int(request.form.get("company_id"), "Empresa")
            except ValueError as exc:
                flash(str(exc), "error")
                return redirect(url_for("manage_companies"))
            row = db.execute("SELECT id, owner_admin_id, is_active FROM companies WHERE id = ?", (company_id,)).fetchone()
            if not row:
                flash("Empresa não encontrada.", "error")
                return redirect(url_for("manage_companies"))
            if not actor_is_master and row["owner_admin_id"] not in actor_owner_ids:
                flash("Sem permissão para alterar empresa de outra comunidade.", "error")
                return redirect(url_for("manage_companies"))
            db.execute("UPDATE companies SET is_active = ?, updated_at = ? WHERE id = ?", (0 if row["is_active"] else 1, now_iso(), company_id))
            db.commit()
            flash("Status da empresa atualizado.", "success")
            return redirect(url_for("manage_companies"))

    companies = fetch_companies(user=actor, include_inactive=True)
    community_admins = fetch_community_admins(user=actor, include_inactive=False)
    show_owner_selector = actor_is_master or len(community_admins) > 1
    return render_template(
        "companies.html",
        companies=companies,
        actor_is_master=actor_is_master,
        community_admins=community_admins,
        show_owner_selector=show_owner_selector,
    )


@app.route("/admin/courses", methods=["GET", "POST"])
@login_required
def manage_courses():
    db = get_db()
    actor = current_user()
    if actor["role"] not in ("admin", "seller"):
        abort(403)
    seller_restricted = actor["role"] == "seller" and not is_manager_user(actor)
    actor_is_master = is_master_user(actor)
    actor_owner_ids = set(get_user_scope_owner_ids(actor))
    if request.method == "POST":
        validate_csrf()
        action = clean_text(request.form.get("action"), 30)

        if action == "create":
            if seller_restricted:
                flash("Vendedores não podem criar cursos. Solicite ao administrador da comunidade.", "error")
                return redirect(url_for("manage_courses"))
            try:
                company_id = parse_int(request.form.get("company_id"), "Empresa")
                default_percent = parse_float(
                    request.form.get("default_commission_percent"),
                    "Comissão padrão",
                    minimum=0,
                    maximum=100,
                )
            except ValueError as exc:
                flash(str(exc), "error")
                return redirect(url_for("manage_courses"))
            company = db.execute(
                "SELECT id, owner_admin_id, is_active FROM companies WHERE id = ?",
                (company_id,),
            ).fetchone()
            if not company or not company["is_active"]:
                flash("Empresa inválida para o curso.", "error")
                return redirect(url_for("manage_courses"))
            if not actor_is_master and company["owner_admin_id"] not in actor_owner_ids:
                flash("Sem permissão para criar curso nesta comunidade.", "error")
                return redirect(url_for("manage_courses"))
            name = clean_text(request.form.get("name"), 120)
            if len(name) < 2:
                flash("Nome do curso inválido.", "error")
                return redirect(url_for("manage_courses"))
            exists = db.execute(
                "SELECT id FROM courses WHERE company_id = ? AND lower(name) = lower(?)",
                (company_id, name),
            ).fetchone()
            if exists:
                flash("Curso já existe nesta empresa.", "error")
                return redirect(url_for("manage_courses"))
            now = now_iso()
            db.execute(
                """
                INSERT INTO courses (company_id, name, default_commission_percent, is_active, created_at, updated_at)
                VALUES (?, ?, ?, 1, ?, ?)
                """,
                (company_id, name, default_percent, now, now),
            )
            db.commit()
            flash("Curso criado.", "success")
            return redirect(url_for("manage_courses"))

        if action == "update":
            try:
                course_id = parse_int(request.form.get("course_id"), "Curso")
                company_id = parse_int(request.form.get("company_id"), "Empresa")
                default_percent = parse_float(
                    request.form.get("default_commission_percent"),
                    "Comissão padrão",
                    minimum=0,
                    maximum=100,
                )
            except ValueError as exc:
                flash(str(exc), "error")
                return redirect(url_for("manage_courses"))
            target = db.execute(
                """
                SELECT c.id, c.company_id, comp.owner_admin_id
                FROM courses c
                INNER JOIN companies comp ON comp.id = c.company_id
                WHERE c.id = ?
                """,
                (course_id,),
            ).fetchone()
            if not target:
                flash("Curso não encontrado.", "error")
                return redirect(url_for("manage_courses"))
            target_company = db.execute(
                "SELECT id, owner_admin_id, is_active FROM companies WHERE id = ?",
                (company_id,),
            ).fetchone()
            if not target_company or not target_company["is_active"]:
                flash("Empresa inválida para o curso.", "error")
                return redirect(url_for("manage_courses"))
            if seller_restricted:
                if not seller_has_course_permission(actor["id"], course_id, "can_edit_course"):
                    flash("Você não tem permissão para editar este curso.", "error")
                    return redirect(url_for("manage_courses"))
                if company_id != target["company_id"]:
                    flash("Vendedor não pode mover curso para outra empresa.", "error")
                    return redirect(url_for("manage_courses"))
            elif not actor_is_master:
                if target["owner_admin_id"] not in actor_owner_ids or target_company["owner_admin_id"] not in actor_owner_ids:
                    flash("Sem permissão para mover/editar curso de outra comunidade.", "error")
                    return redirect(url_for("manage_courses"))
            name = clean_text(request.form.get("name"), 120)
            if len(name) < 2:
                flash("Nome do curso inválido.", "error")
                return redirect(url_for("manage_courses"))
            duplicate = db.execute(
                "SELECT id FROM courses WHERE company_id = ? AND lower(name) = lower(?) AND id <> ?",
                (company_id, name, course_id),
            ).fetchone()
            if duplicate:
                flash("Já existe outro curso com este nome na empresa.", "error")
                return redirect(url_for("manage_courses"))
            db.execute(
                """
                UPDATE courses
                SET company_id = ?, name = ?, default_commission_percent = ?, updated_at = ?
                WHERE id = ?
                """,
                (company_id, name, default_percent, now_iso(), course_id),
            )
            db.commit()
            flash("Curso atualizado.", "success")
            return redirect(url_for("manage_courses"))

        if action == "toggle_active":
            try:
                course_id = parse_int(request.form.get("course_id"), "Curso")
            except ValueError as exc:
                flash(str(exc), "error")
                return redirect(url_for("manage_courses"))
            row = db.execute(
                """
                SELECT c.id, c.is_active, comp.owner_admin_id
                FROM courses c
                INNER JOIN companies comp ON comp.id = c.company_id
                WHERE c.id = ?
                """,
                (course_id,),
            ).fetchone()
            if not row:
                flash("Curso não encontrado.", "error")
                return redirect(url_for("manage_courses"))
            if seller_restricted:
                if not seller_has_course_permission(actor["id"], course_id, "can_edit_course"):
                    flash("Você não tem permissão para editar este curso.", "error")
                    return redirect(url_for("manage_courses"))
            elif not actor_is_master and row["owner_admin_id"] not in actor_owner_ids:
                flash("Sem permissão para alterar curso de outra comunidade.", "error")
                return redirect(url_for("manage_courses"))
            db.execute("UPDATE courses SET is_active = ?, updated_at = ? WHERE id = ?", (0 if row["is_active"] else 1, now_iso(), course_id))
            db.commit()
            flash("Status do curso atualizado.", "success")
            return redirect(url_for("manage_courses"))

    companies = fetch_companies(user=actor, include_inactive=False)
    if seller_restricted:
        courses = fetch_courses(user=actor, include_inactive=True, seller_permission="can_edit_course")
    else:
        courses = fetch_courses(user=actor, include_inactive=True)
    return render_template("courses.html", companies=companies, courses=courses, can_create_course=not seller_restricted)


@app.route("/viewer/course-access", methods=["GET", "POST"])
@roles_required("viewer")
def viewer_course_access():
    user = current_user()
    db = get_db()

    if request.method == "POST":
        validate_csrf()
        try:
            course_id = parse_int(request.form.get("course_id"), "Curso")
        except ValueError as exc:
            flash(str(exc), "error")
            return redirect(url_for("viewer_course_access"))

        note = clean_text(request.form.get("request_note"), 300)
        owner_ids = get_user_scope_owner_ids(user)
        owner_clause, owner_params = build_in_clause("comp.owner_admin_id", owner_ids)
        course = db.execute(
            f"""
            SELECT c.id
            FROM courses c
            INNER JOIN companies comp ON comp.id = c.company_id
            WHERE c.id = ?
              AND c.is_active = 1
              AND comp.is_active = 1
              AND {owner_clause}
            """,
            [course_id, *owner_params],
        ).fetchone()
        if not course:
            flash("Curso inválido para solicitação.", "error")
            return redirect(url_for("viewer_course_access"))

        already_access = db.execute(
            "SELECT id FROM viewer_course_access WHERE viewer_id = ? AND course_id = ?",
            (user["id"], course_id),
        ).fetchone()
        if already_access:
            flash("Você já possui acesso a este curso.", "success")
            return redirect(url_for("viewer_course_access"))

        pending = db.execute(
            """
            SELECT id
            FROM viewer_course_requests
            WHERE viewer_id = ? AND course_id = ? AND status = 'pendente'
            """,
            (user["id"], course_id),
        ).fetchone()
        if pending:
            flash("Já existe uma solicitação pendente para este curso.", "error")
            return redirect(url_for("viewer_course_access"))

        db.execute(
            """
            INSERT INTO viewer_course_requests (
                viewer_id, course_id, request_note, status, requested_at
            ) VALUES (?, ?, ?, 'pendente', ?)
            """,
            (user["id"], course_id, note or None, now_iso()),
        )
        db.commit()
        flash("Solicitação enviada ao administrador.", "success")
        return redirect(url_for("viewer_course_access"))

    today = date.today().isoformat()
    courses = fetch_courses(user=user, include_inactive=False)
    active_accesses = db.execute(
        """
        SELECT a.course_id, c.name AS course_name, comp.name AS company_name, a.is_permanent, a.expires_at
        FROM viewer_course_access a
        INNER JOIN courses c ON c.id = a.course_id
        INNER JOIN companies comp ON comp.id = c.company_id
        WHERE a.viewer_id = ?
          AND (a.is_permanent = 1 OR a.expires_at IS NULL OR a.expires_at >= ?)
        ORDER BY comp.name, c.name
        """,
        (user["id"], today),
    ).fetchall()
    requests = db.execute(
        """
        SELECT r.id, r.course_id, r.request_note, r.status, r.requested_at, r.reviewed_at, r.approval_days, r.is_permanent,
               c.name AS course_name, comp.name AS company_name
        FROM viewer_course_requests r
        INNER JOIN courses c ON c.id = r.course_id
        INNER JOIN companies comp ON comp.id = c.company_id
        WHERE r.viewer_id = ?
        ORDER BY r.requested_at DESC
        """,
        (user["id"],),
    ).fetchall()
    return render_template(
        "viewer_course_access.html",
        courses=courses,
        active_accesses=active_accesses,
        requests=requests,
    )


@app.route("/admin/course-requests", methods=["GET", "POST"])
@login_required
def admin_course_requests():
    db = get_db()
    admin = current_user()
    if not is_admin_or_manager(admin):
        abort(403)
    actor_is_master = is_master_user(admin)
    actor_owner_ids = set(get_user_scope_owner_ids(admin))

    if request.method == "POST":
        validate_csrf()
        action = clean_text(request.form.get("action"), 30)
        try:
            request_id = parse_int(request.form.get("request_id"), "Solicitação")
        except ValueError as exc:
            flash(str(exc), "error")
            return redirect(url_for("admin_course_requests"))

        request_row = db.execute(
            """
            SELECT r.id, r.viewer_id, r.course_id, r.status, comp.owner_admin_id
            FROM viewer_course_requests r
            INNER JOIN courses c ON c.id = r.course_id
            INNER JOIN companies comp ON comp.id = c.company_id
            WHERE id = ?
            """,
            (request_id,),
        ).fetchone()
        if not request_row:
            flash("Solicitação não encontrada.", "error")
            return redirect(url_for("admin_course_requests"))
        if not actor_is_master and request_row["owner_admin_id"] not in actor_owner_ids:
            flash("Sem permissão para analisar solicitação de outra comunidade.", "error")
            return redirect(url_for("admin_course_requests"))
        if request_row["status"] != "pendente":
            flash("Esta solicitação já foi analisada.", "error")
            return redirect(url_for("admin_course_requests"))

        if action == "deny":
            db.execute(
                """
                UPDATE viewer_course_requests
                SET status = 'recusado', reviewed_at = ?, reviewed_by = ?
                WHERE id = ?
                """,
                (now_iso(), admin["id"], request_id),
            )
            db.commit()
            flash("Solicitação recusada.", "success")
            return redirect(url_for("admin_course_requests"))

        if action == "approve":
            permanent = request.form.get("is_permanent") == "1"
            approval_days_raw = (request.form.get("approval_days") or "").strip()
            approval_days = None
            expires_at = None
            if not permanent:
                try:
                    approval_days = parse_int(approval_days_raw, "Dias de acesso", minimum=1)
                except ValueError as exc:
                    flash(str(exc), "error")
                    return redirect(url_for("admin_course_requests"))
                expires_at = (date.today() + timedelta(days=approval_days)).isoformat()

            db.execute(
                """
                INSERT INTO viewer_course_access (
                    viewer_id, course_id, granted_by, granted_at, expires_at, is_permanent
                )
                VALUES (?, ?, ?, ?, ?, ?)
                ON CONFLICT(viewer_id, course_id)
                DO UPDATE SET
                    granted_by = excluded.granted_by,
                    granted_at = excluded.granted_at,
                    expires_at = excluded.expires_at,
                    is_permanent = excluded.is_permanent
                """,
                (
                    request_row["viewer_id"],
                    request_row["course_id"],
                    admin["id"],
                    now_iso(),
                    expires_at,
                    1 if permanent else 0,
                ),
            )
            db.execute(
                """
                UPDATE viewer_course_requests
                SET status = 'aprovado',
                    reviewed_at = ?,
                    reviewed_by = ?,
                    approval_days = ?,
                    is_permanent = ?
                WHERE id = ?
                """,
                (now_iso(), admin["id"], approval_days, 1 if permanent else 0, request_id),
            )
            db.commit()
            flash("Solicitação aprovada.", "success")
            return redirect(url_for("admin_course_requests"))

        flash("Ação inválida para solicitação.", "error")
        return redirect(url_for("admin_course_requests"))

    sql = """
    SELECT r.id, r.status, r.request_note, r.requested_at, r.reviewed_at, r.approval_days, r.is_permanent,
           u.full_name AS viewer_name, u.username AS viewer_username,
           c.name AS course_name, comp.name AS company_name
    FROM viewer_course_requests r
    INNER JOIN users u ON u.id = r.viewer_id
    INNER JOIN courses c ON c.id = r.course_id
    INNER JOIN companies comp ON comp.id = c.company_id
    """
    params = []
    if not actor_is_master:
        owner_clause, owner_params = build_in_clause("comp.owner_admin_id", actor_owner_ids)
        sql += f" WHERE {owner_clause}"
        params.extend(owner_params)
    sql += """
    ORDER BY
      CASE r.status WHEN 'pendente' THEN 0 WHEN 'aprovado' THEN 1 ELSE 2 END,
      r.requested_at DESC
    """
    requests = db.execute(sql, params).fetchall()
    return render_template("admin_course_requests.html", requests=requests)


@app.get("/export/xlsx")
@login_required
def export_xlsx():
    recurring_only = (request.args.get("recurring_only") or "").strip() == "1"
    fallback_route = "dashboard_recorrencia" if recurring_only else "dashboard"
    if Workbook is None:
        flash("Instale openpyxl para exportar xlsx: pip install openpyxl", "error")
        return redirect(url_for(fallback_route))

    user = current_user()
    filters = parse_filters(request.args, user)
    rows = fetch_installment_rows(filters, user, recurring_only=recurring_only)
    if not rows:
        flash("Não há dados para exportar.", "error")
        return redirect(url_for(fallback_route))

    totals, charts = summarize_dashboard(rows)
    wb = Workbook()
    ws = wb.active
    ws.title = "Parcelas Filtradas"
    ws.append(
        [
            "Data Venda",
            "Cliente",
            "Telefone",
            "Empresa",
            "Curso",
            "Vendedor (login)",
            "Pagamento",
            "Modalidade da Comissão",
            "Parcela",
            "Qtd Parcelas",
            "Vencimento",
            "Mês",
            "Valor Parcela",
            "Comissão %",
            "Comissão Parcela",
            "Valor Total Venda",
            "Comissão Total Venda",
            "Status Pagamento",
            "Observações",
        ]
    )

    for row in rows:
        ws.append(
            [
                row["sale_date"],
                row["customer_name"],
                row["customer_phone"] or "",
                row["company_name"],
                row["course_name"],
                row["seller_name"],
                PAYMENT_FORMATS.get(row["payment_format"], row["payment_format"]),
                COMMISSION_PAYMENT_MODES.get(row["commission_payment_mode"], row["commission_payment_mode"]),
                row["installment_number"],
                row["installments_count"],
                row["due_date"],
                row["month_key"],
                float(row["installment_value"]),
                float(row["commission_percent"]),
                float(row["commission_value"]),
                float(row["total_value"]),
                float(row["total_commission_expected"]),
                row.get("display_status_label")
                or (
                    "Pendente de confirmação"
                    if row["installment_status"] == "atrasado" and row["due_date"] > date.today().isoformat()
                    else INSTALLMENT_STATUSES.get(row["installment_status"], row["installment_status"])
                ),
                row["notes"] or "",
            ]
        )

    ws2 = wb.create_sheet("Resumo")
    ws2.append(["Indicador", "Valor"])
    ws2.append(["Parcelas filtradas", totals["count_installments"]])
    ws2.append(["Previsto (valor)", totals["projected_value"]])
    ws2.append(["Previsto (comissão)", totals["projected_commission"]])
    ws2.append(["Confirmado (valor)", totals["confirmed_value"]])
    ws2.append(["Confirmado (comissão)", totals["confirmed_commission"]])
    ws2.append(["Cancelado (valor)", totals["canceled_value"]])
    ws2.append(["Cancelado (comissão)", totals["canceled_commission"]])
    ws2.append(["Pendente (valor)", totals["pending_value"]])
    ws2.append(["Pendente (comissão)", totals["pending_commission"]])
    ws2.append(["Atrasado (valor)", totals["overdue_value"]])
    ws2.append(["Atrasado (comissão)", totals["overdue_commission"]])
    ws2.append(["Exportado em", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])

    ws3 = wb.create_sheet("Projeção Mensal")
    ws3.append(["Mês", "Valor", "Comissão"])
    for idx, month_label in enumerate(charts["monthly"]["labels"]):
        ws3.append([month_label, charts["monthly"]["value"][idx], charts["monthly"]["commission"][idx]])

    ws4 = wb.create_sheet("Projeção Anual")
    ws4.append(["Ano", "Valor", "Comissão"])
    for idx, year_label in enumerate(charts["yearly"]["labels"]):
        ws4.append([year_label, charts["yearly"]["value"][idx], charts["yearly"]["commission"][idx]])

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    file_name = f"comissionamento_filtrado_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return send_file(
        buffer,
        as_attachment=True,
        download_name=file_name,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.errorhandler(400)
def error_400(error):
    message = getattr(error, "description", "Requisição inválida.")
    return render_template("error.html", title="Requisição inválida", message=message), 400


@app.errorhandler(403)
def error_403(_error):
    return render_template("error.html", title="Acesso negado", message="Você não tem permissão para esta ação."), 403


@app.errorhandler(404)
def error_404(_error):
    return render_template("error.html", title="Não encontrado", message="Recurso não encontrado."), 404


init_db()


if __name__ == "__main__":
    port = int(os.environ.get("PORT", "5000"))
    debug = os.environ.get("FLASK_DEBUG", "0") == "1"
    app.run(host="0.0.0.0", port=port, debug=debug)
